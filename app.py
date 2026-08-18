"""Flask Web 应用。

提供以下路由:
- GET  /                          主页 (前端展示)
- POST /api/upload                上传并处理 Excel 文件
- POST /api/process               处理本地文件
- GET  /api/records               获取所有缺陷记录 (JSON)
- GET  /api/records/<row_index>   获取单条记录
- GET  /api/progress/<task_id>    获取处理进度
- GET  /api/image/<path:filepath> 提供图片访问

使用:
    python cli.py serve
    python cli.py serve --port 8080 --debug
"""
from __future__ import annotations

import json
import os
import queue
import subprocess
import sys
from contextlib import closing
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path

from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    render_template,
    request,
    send_file,
    send_from_directory,
)

from core.pipeline import run_pipeline, ProcessPipeline, ProcessConfig

import os
from functools import wraps


# 全局任务状态
TASKS: dict[str, dict] = {}
TASK_LOCK = threading.Lock()
# 每个任务的取消标志 (threading.Event). cancel API 设置 Event, run_task 在每个 progress 回调里检查.
CANCEL_FLAGS: dict[str, threading.Event] = {}

# 任务队列 — 保证同一时间只有一个任务在跑 (避免 LLM 连接冲突)
TASK_QUEUE: queue.Queue = queue.Queue()

def _task_worker():
    """后台 worker, 串行消费任务队列."""
    while True:
        task_id, file_path, output_dir, recognition, task_type, sheet_name = TASK_QUEUE.get()
        # 检查是否在排队期间被取消
        with TASK_LOCK:
            if task_id in CANCEL_FLAGS and CANCEL_FLAGS[task_id].is_set():
                if task_id in TASKS:
                    TASKS[task_id].update({"status": "cancelled", "message": "已取消"})
                CANCEL_FLAGS.pop(task_id, None)
                continue
        try:
            run_task(task_id, file_path, output_dir, recognition, task_type, sheet_name)
        except Exception as e:
            print(f"Task worker error for {task_id}: {e}")
            with TASK_LOCK:
                if task_id in TASKS:
                    TASKS[task_id].update({"status": "failed", "message": str(e)})
        # 任务间重置 LLM 连接 (避免跨任务连接残留)
        try:
            from core.llm_ocr import _reset_conn
            _reset_conn()
        except Exception:
            pass

def _has_dispimg(file_path: str) -> bool:
    """检测文件是否包含 DISPIMG 公式或为 .xls 格式 (需要转换)."""
    path = Path(file_path)
    if path.suffix.lower() == ".xls":
        return True
    # 检查单元格是否有 DISPIMG 公式
    try:
        import zipfile
        with zipfile.ZipFile(path) as zf:
            names_lower = {n.lower() for n in zf.namelist()}
            if "xl/cellimages.xml" in names_lower:
                return True
        from openpyxl import load_workbook
        wb = load_workbook(str(path), data_only=False)
        for ws in wb.worksheets:
            for row in ws.iter_rows(max_row=min(ws.max_row, 10)):
                for cell in row:
                    if cell.value and "DISPIMG" in str(cell.value).upper():
                        return True
    except Exception:
        pass
    return False


def _convert_dispimg(file_path: str) -> str | None:
    """用 convert_DISPIMG_server.py 转换为标准 xlsx, 返回新路径或 None."""
    import subprocess
    output_dir = Path(file_path).parent / "_converted"
    output_dir.mkdir(parents=True, exist_ok=True)
    script = PROJECT_ROOT / "convert_DISPIMG_server.py"
    try:
        result = subprocess.run(
            [sys.executable, str(script), str(file_path), "-o", str(output_dir)],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode != 0:
            print(f"DISPIMG 转换失败: {result.stderr}")
            return None
        # 找到输出文件
        stem = Path(file_path).stem
        for f in output_dir.iterdir():
            if f.suffix.lower() == ".xlsx" and stem in f.name:
                return str(f)
    except Exception as e:
        print(f"DISPIMG 转换异常: {e}")
    return None


# 项目根目录
PROJECT_ROOT = Path(__file__).parent
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output"



# ============= HTTP Basic Auth =============
import base64
import secrets as _secrets
from pathlib import Path as _Path

def _load_credentials():
    user = os.environ.get("ADMIN_USER", "admin")
    pwd = os.environ.get("ADMIN_PASSWORD")
    if not pwd:
        for auth_file in [_Path(__file__).parent / ".auth", _Path("/home/ubuntu/.defect_auth")]:
            if auth_file.exists():
                try:
                    content = auth_file.read_text(encoding="utf-8-sig").strip()
                    if ":" in content:
                        u, p = content.split(":", 1)
                        user = u.strip()
                        pwd = p.strip()
                        break
                except Exception:
                    pass
    if not pwd:
        raise RuntimeError("ADMIN_PASSWORD not set, use env var or .auth file")
    return user, pwd

_AUTH_USER, _AUTH_PASS = _load_credentials()

def _check_auth(auth_header):
    if not auth_header:
        return False
    try:
        scheme, credentials = auth_header.split(" ", 1)
        if scheme.lower() != "basic":
            return False
        decoded = base64.b64decode(credentials).decode("utf-8", errors="ignore")
        if ":" not in decoded:
            return False
        u, p = decoded.split(":", 1)
        return _secrets.compare_digest(u, _AUTH_USER) and _secrets.compare_digest(p, _AUTH_PASS)
    except Exception:
        return False

def _unauthorized():
    return Response(
        "<!doctype html><html><head><meta charset=utf-8>" +
        "<title>需要登录 - 钢材缺陷图像知识库</title>" +
        "<style>body{font-family:sans-serif;display:flex;align-items:center;" +
        "justify-content:center;height:100vh;margin:0;background:#1a202c;color:#fff;}" +
        ".box{text-align:center;padding:2rem;}" +
        "h1{margin:0 0 1rem;font-size:1.5rem;}" +
        "p{opacity:0.7;}</style></head><body>" +
        "<div class=box><h1>钢材缺陷图像知识库 - 需要登录</h1>" +
        "<p>Please use account and password (browser will prompt)</p></div></body></html>",
        status=401,
        headers={"WWW-Authenticate": "Basic realm=Steel-Defect-Knowledge-Base"}
    )

def auth_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.path == "/api/health":
            return f(*args, **kwargs)
        if not _check_auth(request.headers.get("Authorization")):
            return _unauthorized()
        return f(*args, **kwargs)
    return decorated

# ============= /HTTP Basic Auth =============


def _load_cscan_from_db(conn, task_id: str) -> list[dict]:
    """从 SQLite cscan_records 读 records, 反序列化 缺陷表格 JSON."""
    import json as _json
    rows = conn.execute(
        "SELECT * FROM cscan_records WHERE task_id = ? ORDER BY row_index",
        (task_id,),
    ).fetchall()
    # PRAGMA table_info 返回 (cid, name, type, notnull, dflt_value, pk), name 在 index 1
    cols = [d[1] for d in conn.execute("PRAGMA table_info(cscan_records)").fetchall()]
    results = []
    for row in rows:
        d = dict(zip(cols, row))
        for fld in ("缺陷表格_F", "缺陷表格_G"):
            try:
                d[fld] = _json.loads(d.get(fld) or "[]")
            except Exception:
                d[fld] = []
        try:
            d["warnings"] = _json.loads(d.get("warnings") or "[]")
        except Exception:
            d["warnings"] = []
        results.append(d)
    return results


def _load_zhongban_from_db(output_dir: Path) -> list[dict] | None:
    """从 SQLite 读取 zhongban 记录, 返回 JSON 兼容格式. 无数据返回 None."""
    import sqlite3 as _sqlite3
    db_path = PROJECT_ROOT / "data" / "defect_map.db"
    if not db_path.exists():
        return None
    try:
        with closing(_sqlite3.connect(str(db_path))) as conn:
            conn.row_factory = _sqlite3.Row
            # 按 output_dir 找 task (DB 存的是绝对路径)
            od_abs = str(output_dir.resolve())
            task = conn.execute(
                "SELECT id FROM tasks WHERE output_dir = ? LIMIT 1",
                (od_abs,),
            ).fetchone()
            if not task:
                # 也尝试匹配末尾 (兼容不同 base path)
                task = conn.execute(
                    "SELECT id FROM tasks WHERE output_dir LIKE ? LIMIT 1",
                    (f"%{output_dir.name}",),
                ).fetchone()
            if not task:
                return None
            task_pk = task[0]

            # 读所有 records
            db_records = conn.execute(
                "SELECT * FROM records WHERE task_id = ? ORDER BY row_index",
                (task_pk,),
            ).fetchall()
            if not db_records:
                return None

            result = []
            for rec in db_records:
                r = dict(rec)
                ri = r["row_index"]
                # 基础字段
                out = {
                    "row_index": ri,
                    "序号": str(r.get("sequence", "") or ""),
                    "生产厂": str(r.get("factory", "") or ""),
                    "钢板号": str(r.get("plate_no", "") or ""),
                    "钢种": str(r.get("steel_grade", "") or ""),
                    "类别": str(r.get("category", "") or ""),
                    "缺陷分析": str(r.get("defect_analysis", "") or ""),
                    "缺陷图谱": "", "缺陷照片": "",
                }
                # images + views + ocr
                images = conn.execute(
                    "SELECT * FROM images WHERE record_id = ? ORDER BY image_index",
                    (r["id"],),
                ).fetchall()
                img_params = {}
                for idx, img in enumerate(images, 1):
                    img_d = dict(img)
                    out[f"图-{idx}"] = img_d.get("file_path", "")
                    out[f"图-{idx}_format"] = img_d.get("image_format", "")
                    # views — 直接用 DB 里存的 view_label 作为 key
                    views = conn.execute(
                        "SELECT * FROM views WHERE image_id = ?",
                        (img_d["id"],),
                    ).fetchall()
                    for v in views:
                        vl = v["view_label"]
                        if isinstance(vl, str) and vl:
                            # view_label 已含图序号后缀 (如 "俯视图-1"), 直接用作 key
                            out[vl] = v["file_path"]
                    # ocr
                    ocr = conn.execute(
                        "SELECT * FROM ocr_results WHERE image_id = ?",
                        (img_d["id"],),
                    ).fetchone()
                    if ocr:
                        od = dict(ocr)
                        for k in ("material_size", "defect_center_x", "defect_center_y",
                                  "defect_length", "defect_width", "defect_depth"):
                            v = str(od.get(k, "") or "")
                            if v:
                                img_params[_ocr_key(k)] = v
                # 合并 缺陷数据
                out["缺陷数据"] = img_params
                out["warnings"] = []
                result.append(out)
            return result if result else None
    except Exception as e:
        print(f"load zhongban from DB failed: {e}")
        return None


def _ocr_key(db_col: str) -> str:
    """ocr_results 列名 → 缺陷数据 key."""
    return {
        "material_size": "材料尺寸",
        "defect_center_x": "缺陷中心X",
        "defect_center_y": "缺陷中心Y",
        "defect_length": "缺陷长度",
        "defect_width": "缺陷宽度",
        "defect_depth": "缺陷深度",
    }.get(db_col, db_col)


def _db_sync_zhongban_defect(output_dir: Path, row_index: int, defects: dict):
    """同步 zhongban 缺陷数据到 SQLite ocr_results 表."""
    import sqlite3 as _sqlite3
    db_path = PROJECT_ROOT / "data" / "defect_map.db"
    if not db_path.exists():
        return
    try:
        with closing(_sqlite3.connect(str(db_path))) as conn:
            od_abs = str(output_dir.resolve())
            ocr_row = conn.execute("""
                SELECT ocr.id FROM ocr_results ocr
                JOIN images i ON i.id = ocr.image_id
                JOIN records r ON r.id = i.record_id
                JOIN tasks t ON t.id = r.task_id
                WHERE (t.output_dir = ? OR t.output_dir LIKE ?)
                AND r.row_index = ?
                ORDER BY i.image_index LIMIT 1
            """, (od_abs, f"%{output_dir.name}", row_index)).fetchone()
            if ocr_row:
                conn.execute("""
                    UPDATE ocr_results SET
                        material_size = ?,
                        defect_center_x = ?,
                        defect_center_y = ?,
                        defect_length = ?,
                        defect_width = ?,
                        defect_depth = ?
                    WHERE id = ?
                """, (
                    defects.get("材料尺寸", ""),
                    defects.get("缺陷中心X", ""),
                    defects.get("缺陷中心Y", ""),
                    defects.get("缺陷长度", ""),
                    defects.get("缺陷宽度", ""),
                    defects.get("缺陷深度", ""),
                    ocr_row[0],
                ))
                conn.commit()
    except Exception as e:
        print(f"DB sync zhongban defect failed: {e}")


def _db_sync_cscan_record(task_id: str, row_index: int, record: dict):
    """同步单条 cscan 记录到 SQLite (编辑后保持 JSON 和 DB 一致)."""
    import sqlite3 as _sqlite3
    import json as _json
    db_path = PROJECT_ROOT / "data" / "defect_map.db"
    if not db_path.exists():
        return
    try:
        with closing(_sqlite3.connect(str(db_path))) as conn:
            # 序列化 JSON 字段
            ft = _json.dumps(record.get("缺陷表格_F") or [], ensure_ascii=False)
            gt = _json.dumps(record.get("缺陷表格_G") or [], ensure_ascii=False)
            warns = _json.dumps(record.get("warnings") or [], ensure_ascii=False)
            conn.execute(
                """UPDATE cscan_records SET
                   \"序号\"=?, \"生产厂\"=?, \"钢板号\"=?, \"钢种\"=?,
                   \"类别\"=?, \"缺陷分析\"=?,
                   缺陷表格_F=?, 缺陷表格_G=?, warnings=?
                   WHERE task_id=? AND row_index=?""",
                (
                    str(record.get("序号", "")),
                    str(record.get("生产厂", "")),
                    str(record.get("钢板号", "")),
                    str(record.get("钢种", "")),
                    str(record.get("类别", "")),
                    str(record.get("缺陷分析", "")),
                    ft, gt, warns,
                    task_id, row_index,
                ),
            )
            conn.commit()
    except Exception as e:
        print(f"DB sync cscan_record failed: {e}")


def create_app() -> Flask:
    """创建 Flask 应用。"""
    app = Flask(
        __name__,
        template_folder=str(PROJECT_ROOT / "templates"),
        static_folder=str(PROJECT_ROOT / "static"),
    )
    app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200MB 上限

    register_routes(app)

    # 启动任务队列 worker
    worker = threading.Thread(target=_task_worker, daemon=True)
    worker.start()

    return app


def register_routes(app: Flask):
    """注册路由。"""

    @app.route("/")
    @auth_required
    def index():
        """主页。"""
        return render_template("index.html")

    @app.route("/api/upload", methods=["POST"])
    @auth_required
    def api_upload():
        """上传并处理 Excel 文件。"""
        if "file" not in request.files:
            return jsonify({"error": "未提供文件"}), 400

        file = request.files["file"]
        if not file.filename:
            return jsonify({"error": "文件名为空"}), 400

        # 检查扩展名
        ext = Path(file.filename).suffix.lower()
        if ext not in (".xls", ".xlsx"):
            return jsonify({"error": f"不支持的格式: {ext}, 仅支持 .xls/.xlsx"}), 400

        # 保存到 input 目录
        input_dir = PROJECT_ROOT / "input"
        input_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = Path(file.filename).stem
        saved_path = input_dir / f"{safe_name}_{timestamp}{ext}"
        file.save(str(saved_path))

        # 检测 DISPIMG / .xls → 自动转换为标准 xlsx
        if _has_dispimg(str(saved_path)):
            converted_path = _convert_dispimg(str(saved_path))
            if converted_path:
                saved_path = converted_path
                ext = ".xlsx"

        # 后台线程处理
        recognition = request.form.get("recognition", "ocr")
        task_type = request.form.get("task_type", "zhongban").lower()
        if task_type not in ("zhongban", "cscan", "kuanhouban"):
            task_type = "zhongban"

        task_ids = []
        if task_type in ("cscan", "kuanhouban"):
            # 模板二: 发现所有日期 sheet / 模板三: 发现所有有数据的 sheet (跳过 Sheet1)
            if task_type == "kuanhouban":
                from core.cscan_ocr import find_kuanhouban_sheets
                sheets = find_kuanhouban_sheets(str(saved_path))
            else:
                from core.cscan_ocr import find_cscan_sheets
                sheets = find_cscan_sheets(str(saved_path))
            if not sheets:
                return jsonify({
                    "error": "未找到可处理的 sheet。"
                             "模板二需要日期格式 sheet (如 5.1)，模板三检查 Sheet2/Sheet3 等。"
                }), 400
            for item in sheets:
                if isinstance(item, tuple):
                    s, label = item  # kuanhouban: (sheet_name, date_label)
                else:
                    s = label = item  # cscan: plain string
                tid = str(uuid.uuid4())[:8]
                od = DEFAULT_OUTPUT_DIR / tid
                with TASK_LOCK:
                    TASKS[tid] = {
                        "status": "pending",
                        "progress": 0.0,
                        "stage": "init",
                        "message": f"等待处理 sheet {label}...",
                        "file": str(saved_path),
                        "output_dir": str(od),
                        "created_at": time.time(),
                        "task_type": task_type,
                        "sheet_name": label,
                    }
                CANCEL_FLAGS[tid] = threading.Event()
                TASK_QUEUE.put((tid, str(saved_path), str(od), recognition, task_type, s))
                task_ids.append(tid)
        else:
            # 模板一 / 模板三: 单个任务
            task_id = str(uuid.uuid4())[:8]
            output_dir = DEFAULT_OUTPUT_DIR / task_id
            sheet = ""  # kuanhouban/cscan 自动检测, zhongban 不需要
            with TASK_LOCK:
                TASKS[task_id] = {
                    "status": "pending",
                    "progress": 0.0,
                    "stage": "init",
                    "message": "排队等待...",
                    "file": str(saved_path),
                    "output_dir": str(output_dir),
                    "created_at": time.time(),
                    "task_type": task_type,
                }
            CANCEL_FLAGS[task_id] = threading.Event()
            TASK_QUEUE.put((task_id, str(saved_path), str(output_dir), recognition, task_type, sheet))
            task_ids.append(task_id)

        return jsonify({
            "task_ids": task_ids,
            "sheet_count": len(task_ids),
            "status": "processing",
        })

    @app.route("/api/process", methods=["POST"])
    @auth_required
    def api_process():
        """处理已存在的本地文件路径。"""
        data = request.get_json()
        file_path = data.get("file_path") if data else None
        if not file_path:
            return jsonify({"error": "未提供 file_path"}), 400

        if not Path(file_path).exists():
            return jsonify({"error": f"文件不存在: {file_path}"}), 404

        # 检查扩展名 - 只支持 .xlsx
        ext = Path(file_path).suffix.lower()
        if ext == ".xls":
            return jsonify({
                "error": "不支持 .xls 格式! 请先用 WPS/Excel 打开并另存为 .xlsx 后再处理。"
            }), 400
        if ext != ".xlsx":
            return jsonify({"error": f"不支持的格式: {ext}, 仅支持 .xlsx"}), 400

        task_id = str(uuid.uuid4())[:8]
        output_dir = DEFAULT_OUTPUT_DIR / task_id
        with TASK_LOCK:
            TASKS[task_id] = {
                "status": "pending",
                "progress": 0.0,
                "stage": "init",
                "message": "等待处理...",
                "file": file_path,
                "output_dir": str(output_dir),
                "created_at": time.time(),
            }

        recognition = data.get("recognition", "ocr")

        TASK_QUEUE.put((task_id, file_path, str(output_dir), recognition, "zhongban", ""))

        return jsonify({"task_id": task_id, "status": "queued"})

    @app.route("/api/progress/<task_id>")
    @auth_required
    def api_progress(task_id: str):
        """查询任务进度。"""
        with TASK_LOCK:
            task = TASKS.get(task_id)
            if not task:
                return jsonify({"error": "任务不存在"}), 404
            return jsonify(task)

    @app.route("/api/records/<task_id>")
    @auth_required
    def api_records(task_id: str):
        """获取任务的缺陷记录。SQLite 优先, JSON 兜底。"""
        with TASK_LOCK:
            task = TASKS.get(task_id)

        if task:
            output_dir = Path(task["output_dir"])
            status = task["status"]
        else:
            output_dir = DEFAULT_OUTPUT_DIR / task_id
            if not output_dir.exists():
                return jsonify({"error": "任务不存在"}), 404
            status = "completed"

        # 1. SQLite 优先
        records = _load_zhongban_from_db(output_dir)
        # 2. JSON 兜底
        if records is None:
            json_path = output_dir / "defect_records.json"
            if not json_path.exists():
                return jsonify({"error": "记录尚未生成", "status": status}), 404
            with open(json_path, "r", encoding="utf-8") as f:
                records = json.load(f)

        return jsonify(
            {
                "task_id": task_id,
                "status": status,
                "count": len(records),
                "records": records,
            }
        )

    @app.route("/api/download/<task_id>")
    @auth_required
    def api_download_zip(task_id: str):
        """打包下载: ZIP 包含 JSON + images/ + views/."""
        import io, zipfile as _zipfile

        with TASK_LOCK:
            task = TASKS.get(task_id)
        if task:
            output_dir = Path(task["output_dir"])
        else:
            output_dir = DEFAULT_OUTPUT_DIR / task_id
        if not output_dir.exists():
            abort(404)

        # 确定 JSON 文件
        cscan_jf = output_dir / "cscan_records.json"
        zhongban_jf = output_dir / "defect_records.json"
        json_file = cscan_jf if cscan_jf.exists() else zhongban_jf
        if not json_file.exists():
            abort(404)

        # 创建内存 ZIP
        buf = io.BytesIO()
        with _zipfile.ZipFile(buf, "w", _zipfile.ZIP_DEFLATED) as zf:
            # 1. JSON
            zf.write(str(json_file), json_file.name)

            # 2. images/
            images_dir = output_dir / "images"
            if images_dir.exists():
                for f in sorted(images_dir.iterdir()):
                    if f.is_file():
                        zf.write(str(f), f"images/{f.name}")

            # 3. views/ (zhongban)
            views_dir = output_dir / "views"
            if views_dir.exists():
                for f in sorted(views_dir.rglob("*")):
                    if f.is_file():
                        arcname = str(f.relative_to(output_dir))
                        zf.write(str(f), arcname)

        buf.seek(0)
        ti = json.load(open(output_dir / "task_info.json")) if (output_dir / "task_info.json").exists() else {}
        label = ti.get("sheet_name", "") or task_id
        return send_file(buf, mimetype="application/zip", as_attachment=True,
                         download_name=f"{label}.zip")

    @app.route("/api/image/<task_id>/<path:filepath>")
    @auth_required
    def api_image(task_id: str, filepath: str):
        """提供图片访问。"""
        with TASK_LOCK:
            task = TASKS.get(task_id)
        if task:
            output_dir = Path(task["output_dir"])
        else:
            output_dir = DEFAULT_OUTPUT_DIR / task_id
            if not output_dir.exists():
                abort(404)

        # 安全检查：防止路径穿越
        try:
            img_path = (output_dir / filepath).resolve()
            output_dir_resolved = output_dir.resolve()
            if not str(img_path).startswith(str(output_dir_resolved)):
                abort(403)
            if not img_path.exists():
                abort(404)
            return send_from_directory(str(img_path.parent), img_path.name)
        except Exception:
            abort(404)

    @app.route("/api/list")
    @auth_required
    def api_list():
        """列出所有任务。

        数据源:
        1. 内存中的 TASKS dict (本次进程内的任务, 含实时进度)
        2. 磁盘 output/<task_id>/defect_records.json (历史任务, 重启后仍可见)
        """
        tasks = []
        seen = set()

        # 1. 内存中的活跃任务
        with TASK_LOCK:
            for tid, t in TASKS.items():
                # 尝试从输出目录读取记录条数
                count = None
                output_dir = Path(t.get("output_dir", "")) if t.get("output_dir") else None
                if output_dir and output_dir.exists():
                    cscan_path = output_dir / "cscan_records.json"
                    json_path = cscan_path if cscan_path.exists() else (output_dir / "defect_records.json")
                    if json_path.exists():
                        try:
                            with open(json_path, "r", encoding="utf-8") as f:
                                data = json.load(f)
                            records = data.get("records", data) if isinstance(data, dict) else data
                            count = len(records) if isinstance(records, list) else 0
                        except Exception:
                            pass
                task_entry = {
                    "task_id": tid,
                    "status": t["status"],
                    "stage": t.get("stage", ""),
                    "progress": t.get("progress", 0),
                    "file": Path(t["file"]).name if t.get("file") else "",
                    "created_at": t.get("created_at", 0),
                    "task_type": t.get("task_type", "zhongban"),
                    "sheet_name": t.get("sheet_name", ""),
                    "source": "memory",
                }
                if count is not None:
                    task_entry["count"] = count
                tasks.append(task_entry)
                seen.add(tid)

        # 2. 磁盘上的历史任务
        for task_dir in DEFAULT_OUTPUT_DIR.iterdir():
            if not task_dir.is_dir():
                continue
            tid = task_dir.name
            if tid in seen:
                continue
            json_path = task_dir / "defect_records.json"
            # cscan 任务用 cscan_records.json, 兼容两种
            cscan_path = task_dir / "cscan_records.json"
            json_path_eff = cscan_path if cscan_path.exists() else json_path
            if not json_path_eff.exists():
                continue
            # 读取元数据
            try:
                with open(json_path_eff, "r", encoding="utf-8") as f:
                    data = json.load(f)
                # 文件可能是 list 也可能是 dict (带 count/records 字段)
                records = data.get("records", data) if isinstance(data, dict) else data
                count = len(records) if isinstance(records, list) else 0
            except Exception:
                count = 0
            # 从 task_info.json 读元数据
            ti_data = {}
            ti_path = task_dir / "task_info.json"
            if ti_path.exists():
                try:
                    with open(ti_path) as tf:
                        ti_data = json.load(tf)
                except Exception:
                    pass
            detected_type = ti_data.get("task_type") or ("cscan" if cscan_path.exists() else "zhongban")

            # 从数据库查文件名 (按 output_dir 匹配 tasks 表)
            file_name = ti_data.get("source_file", "")
            if not file_name:
                try:
                    import sqlite3
                    db_path = PROJECT_ROOT / "data" / "defect_map.db"
                    with closing(sqlite3.connect(str(db_path))) as conn:
                        row = conn.execute(
                            """SELECT f.filename FROM tasks t
                               LEFT JOIN files f ON f.id = t.file_id
                               WHERE t.output_dir = ?
                               LIMIT 1""",
                            (str(task_dir),),
                        ).fetchone()
                        if row and row[0]:
                            file_name = row[0]
                except Exception:
                    pass
            if not file_name:
                file_name = f"任务 {tid} ({count} 条)"

            sheet_name = ti_data.get("sheet_name", "")

            # 创建时间 = 目录 mtime
            created_at = task_dir.stat().st_mtime
            tasks.append(
                {
                    "task_id": tid,
                    "status": "completed",
                    "stage": "database",
                    "progress": 1.0,
                    "file": file_name,
                    "created_at": created_at,
                    "count": count,
                    "task_type": detected_type,
                    "sheet_name": sheet_name,
                    "source": "disk",
                }
            )

        tasks.sort(key=lambda x: x["created_at"], reverse=True)
        return jsonify({"tasks": tasks})

    @app.route("/api/cancel/<task_id>", methods=["POST"])
    @auth_required
    def api_cancel(task_id: str):
        """取消正在跑或排队中的任务。"""
        with TASK_LOCK:
            event = CANCEL_FLAGS.get(task_id)
            if event is None:
                task = TASKS.get(task_id)
                if task and task.get("status") in ("completed", "failed", "cancelled"):
                    return jsonify({"ok": True, "message": f"任务已是 {task.get('status')} 状态, 无需取消"}), 200
                return jsonify({"ok": False, "error": f"任务 {task_id} 不在跑"}), 404
            event.set()
            # 如果还在排队, 直接标记取消 (worker 会跳过)
            if task_id in TASKS and TASKS[task_id].get("status") == "pending":
                TASKS[task_id].update({"status": "cancelled", "message": "已取消"})
        return jsonify({"ok": True, "message": f"已取消 {task_id}"}), 200

    @app.route("/api/task/<task_id>", methods=["DELETE"])
    @auth_required
    def api_delete_task(task_id: str):
        """删除任务: 清理内存状态、输出目录、数据库记录."""
        import shutil
        import sqlite3 as _sqlite3

        # 1. 从内存中移除
        with TASK_LOCK:
            TASKS.pop(task_id, None)
            CANCEL_FLAGS.pop(task_id, None)

        # 2. 删除磁盘输出目录
        output_dir = DEFAULT_OUTPUT_DIR / task_id
        if output_dir.exists():
            try:
                shutil.rmtree(str(output_dir))
            except Exception as e:
                return jsonify({"ok": False, "error": f"删除目录失败: {e}"}), 500

        # 3. 删除数据库记录 (tasks 表按 output_dir 匹配)
        db_path = PROJECT_ROOT / "data" / "defect_map.db"
        if db_path.exists():
            try:
                with closing(_sqlite3.connect(str(db_path))) as conn:
                    conn.execute("PRAGMA foreign_keys = ON")
                    conn.execute(
                        "DELETE FROM cscan_records WHERE task_id = ?", (task_id,)
                    )
                    conn.execute(
                        "DELETE FROM tasks WHERE output_dir = ?", (str(output_dir),)
                    )
            except Exception as e:
                # DB 清理失败不阻塞，目录已删
                print(f"清理数据库失败: {e}")

        return jsonify({"ok": True, "task_id": task_id})

    @app.route("/api/summary/<task_type>")
    @auth_required
    def api_summary(task_type: str):
        """厂汇总: 合并该类型所有已完成任务的记录."""
        if task_type not in ("zhongban", "cscan", "kuanhouban"):
            return jsonify({"error": "无效类型"}), 400

        all_records = []
        total = 0

        for task_dir in DEFAULT_OUTPUT_DIR.iterdir():
            if not task_dir.is_dir():
                continue
            # 查 task_info.json 确定类型
            ti_path = task_dir / "task_info.json"
            if not ti_path.exists():
                continue
            try:
                with open(ti_path) as f:
                    ti = json.load(f)
            except Exception:
                continue
            if ti.get("task_type", "zhongban") != task_type:
                continue

            # 读记录
            if task_type == "zhongban":
                jf = task_dir / "defect_records.json"
                if jf.exists():
                    with open(jf) as f:
                        data = json.load(f)
                    recs = data if isinstance(data, list) else data.get("records", [])
                else:
                    recs = _load_zhongban_from_db(task_dir) or []
            else:
                jf = task_dir / "cscan_records.json"
                if jf.exists():
                    with open(jf) as f:
                        data = json.load(f)
                    recs = data.get("records", data) if isinstance(data, dict) else data
                else:
                    recs = []

            for r in recs:
                r["_task_id"] = task_dir.name
                r["_sheet"] = ti.get("sheet_name", "")
            all_records.extend(recs)
            total += len(recs)

        return jsonify({
            "task_type": task_type,
            "task_count": total,
            "records": all_records,
        })

    @app.route("/api/cscan_records/<task_id>")
    @auth_required
    def api_cscan_records(task_id: str):
        """读 cscan_records (中厚板卷厂 schema). 从 SQLite 优先, fallback JSON."""
        try:
            # 优先从内存里的 TASKS (实时) 拿 output_dir
            output_dir = None
            with TASK_LOCK:
                t = TASKS.get(task_id)
                if t:
                    output_dir = Path(t.get("output_dir", ""))
            # 兜底: 磁盘路径
            if not output_dir or not output_dir.exists():
                output_dir = DEFAULT_OUTPUT_DIR / task_id
            if not output_dir.exists():
                return jsonify({"error": f"任务 {task_id} 不存在"}), 404

            # 优先从 SQLite 读
            import sqlite3
            db_path = PROJECT_ROOT / "data" / "defect_map.db"
            db_records = None
            if db_path.exists():
                try:
                    conn = sqlite3.connect(str(db_path))
                    try:
                        db_records = _load_cscan_from_db(conn, task_id)
                    finally:
                        conn.close()
                except Exception:
                    pass  # DB 表不存在/损坏 → 回退到 JSON
            if db_records:
                return jsonify({"task_id": task_id, "count": len(db_records), "records": db_records})

            # Fallback 从 JSON 读
            json_path = output_dir / "cscan_records.json"
            if not json_path.exists():
                return jsonify({"error": f"任务 {task_id} 没有 cscan_records"}), 404
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return jsonify({"task_id": task_id, **data})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/cscan_records_xlsx/<task_id>", methods=["GET", "POST"])
    @auth_required
    def api_cscan_xlsx(task_id: str):
        """提供 cscan_records.xlsx 下载.

        GET  → 全量 (无 filter)
        POST → body: {"row_indexes": [8,9,11,...]} 按 filter 导出
        """
        import json as _json
        from core.cscan_merger import save_excel
        with TASK_LOCK:
            task = TASKS.get(task_id)
        output_dir = Path(task["output_dir"]) if task else DEFAULT_OUTPUT_DIR / task_id
        json_path = output_dir / "cscan_records.json"
        if not json_path.exists():
            abort(404)
        with open(json_path, "r", encoding="utf-8") as f:
            all_data = _json.load(f)
        all_records = all_data.get("records", [])
        if not all_records:
            abort(404)

        # 如果前端发了 row_indexes (filter+sort 后的行), 只导出这些
        if request.method == "POST":
            body = request.get_json(silent=True) or {}
            wanted = set(body.get("row_indexes") or [])
            if wanted:
                all_records = [r for r in all_records if r.get("row_index") in wanted]

        xlsx_path = output_dir / "cscan_records.xlsx"
        save_excel(all_records, output_dir)

        # 从 task_info 获取源文件名作为下载名
        source_name = task_id
        ti_path = output_dir / "task_info.json"
        if ti_path.exists():
            try:
                with open(ti_path, "r", encoding="utf-8") as tf:
                    ti = _json.load(tf)
                src = ti.get("source_file", "")
                if src:
                    source_name = Path(src).stem
            except Exception:
                pass
        if not source_name or source_name == task_id:
            with TASK_LOCK:
                t = TASKS.get(task_id)
            if t and t.get("file"):
                source_name = Path(t["file"]).stem

        return send_file(str(xlsx_path), as_attachment=True,
                         download_name=f"{source_name}_CScan记录.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


    @app.route("/api/cscan_records/<task_id>/<int:row_index>", methods=["PUT"])
    @auth_required
    def api_update_cscan_record(task_id: str, row_index: int):
        """更新 cscan 记录的缺陷表格数据."""
        data = request.get_json(silent=True) or {}
        with TASK_LOCK:
            task = TASKS.get(task_id)
        output_dir = Path(task["output_dir"]) if task else DEFAULT_OUTPUT_DIR / task_id
        json_path = output_dir / "cscan_records.json"
        if not json_path.exists():
            return jsonify({"error": "cscan_records not found"}), 404
        with open(json_path, "r", encoding="utf-8") as f:
            all_data = json.load(f)
        records = all_data.get("records", [])
        updated = None
        for r in records:
            if r.get("row_index") == row_index:
                for k, v in data.items():
                    r[k] = v
                updated = r
                break
        if updated is None:
            return jsonify({"error": f"row_index {row_index} not found"}), 404
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(all_data, f, ensure_ascii=False, indent=2)
        # 重新生成 xlsx
        from core.cscan_merger import save_excel
        save_excel(records, output_dir)
        # 同步更新 SQLite
        _db_sync_cscan_record(task_id, row_index, updated)
        return jsonify({"ok": True, "row_index": row_index})

    @app.route("/api/records/<task_id>/<int:row_index>", methods=["POST", "PUT"])
    @auth_required
    def api_update_record(task_id: str, row_index: int):
        """更新某条记录的缺陷数据 (用户手动编辑)。"""
        # 找输出目录
        with TASK_LOCK:
            task = TASKS.get(task_id)
        if task:
            output_dir = Path(task["output_dir"])
        else:
            output_dir = DEFAULT_OUTPUT_DIR / task_id
        if not output_dir.exists():
            return jsonify({"error": "任务不存在"}), 404

        json_path = output_dir / "defect_records.json"
        if not json_path.exists():
            return jsonify({"error": "记录尚未生成"}), 404

        # 读现有数据
        with open(json_path, "r", encoding="utf-8") as f:
            records = json.load(f)

        # 找 row
        target = None
        for r in records:
            if r.get("row_index") == row_index:
                target = r
                break
        if target is None:
            return jsonify({"error": f"未找到 row_index={row_index}"}), 404

        # 取 POST body
        data = request.get_json() or {}
        defects = data.get("缺陷数据", {})
        if not isinstance(defects, dict):
            return jsonify({"error": "缺陷数据必须是对象"}), 400

        # 更新该 row 的 缺陷数据
        target["缺陷数据"] = defects

        # 写回
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)

        # 重新生成 Excel, 让编辑反映到下载的 excel
        try:
            from core.data_merger import DataMerger
            merger = DataMerger(output_dir)
            merger.save_excel(records)
        except Exception as e:
            # Excel 生成失败不阻塞保存
            print(f"重新生成 Excel 失败: {e}")

        # 同步到 SQLite
        _db_sync_zhongban_defect(output_dir, row_index, defects)

        return jsonify({
            "success": True,
            "row_index": row_index,
            "缺陷数据": defects,
        })

    @app.route("/api/health")
    @auth_required
    def api_health():
        """健康检查。"""
        return jsonify({"status": "ok", "time": datetime.now().isoformat()})


def run_task(
    task_id: str,
    file_path: str,
    output_dir: str,
    recognition: str = "ocr",
    task_type: str = "zhongban",
    sheet_name: str = "",
):
    """后台执行处理任务。sheet_name 用于 cscan 模板二, 指定要处理的日期 sheet."""

    # 把 sheet_name 存到 task_info.json, 供重启后恢复
    import json as _json
    ti = {"source_file": Path(file_path).name, "task_type": task_type}
    if sheet_name:
        ti["sheet_name"] = sheet_name
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    with open(Path(output_dir) / "task_info.json", "w", encoding="utf-8") as tf:
        _json.dump(ti, tf, ensure_ascii=False)

    # 复用已有的取消标志 (可能已由队列创建), 否则新建
    with TASK_LOCK:
        cancel_event = CANCEL_FLAGS.get(task_id) or threading.Event()
        CANCEL_FLAGS[task_id] = cancel_event

    try:
        def progress(stage, percent, message):
            # 检查是否被用户取消
            if cancel_event.is_set():
                raise InterruptedError("任务被用户取消")
            with TASK_LOCK:
                if task_id in TASKS:
                    TASKS[task_id].update(
                        {
                            "stage": stage,
                            "progress": percent,
                            "message": message,
                            "status": "processing",
                        }
                    )

        config = ProcessConfig(
            file_path=file_path,
            output_dir=output_dir,
            enable_ocr=(task_type == "zhongban"),
            enable_split=(task_type == "zhongban"),
            task_type=task_type,
            recognition=recognition,
            sheet_name=sheet_name,
        )
        pipeline = ProcessPipeline(config)
        pipeline.cancel_event = cancel_event
        try:
            result = pipeline.run(progress_callback=progress)
        except InterruptedError:
            with TASK_LOCK:
                if task_id in TASKS:
                    TASKS[task_id].update(
                        {
                            "status": "cancelled",
                            "message": "已取消",
                        }
                    )
            return

        with TASK_LOCK:
            if task_id in TASKS:
                TASKS[task_id].update(
                    {
                        "status": "completed" if result.success else "failed",
                        "progress": 1.0,
                        "message": (
                            "处理完成"
                            if result.success
                            else f"失败: {result.error}"
                        ),
                        "stats": result.stats,
                        "json_path": result.json_path,
                        "excel_path": result.excel_path,
                    }
                )
    finally:
        # 清理
        with TASK_LOCK:
            CANCEL_FLAGS.pop(task_id, None)


# ============================================================================
# 入口
# ============================================================================
app = create_app()


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)