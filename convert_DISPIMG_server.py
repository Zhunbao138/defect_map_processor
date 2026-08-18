"""将 WPS 的 ``DISPIMG`` 图片转换为标准 XLSX 浮动图片。

脚本处理 WPS 的非标准图片函数，以及旧式 XLS 中的原生图片对象，不依赖
WPS/Excel。输入可以是 ``.xls`` 或 ``.xlsx``，结果默认保存到输入目录下的
``finishexcel`` 文件夹。

依赖：
    pip install xlrd olefile openpyxl Pillow

示例：
    python convert_DISPIMG_server.py
    python convert_DISPIMG_server.py "D:/work/book.xls"
    python convert_DISPIMG_server.py "D:/work" --output-dir "D:/work/out"
"""

from __future__ import annotations

import argparse
import io
import posixpath
import re
import struct
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET


EMU_PER_PX = 9525
MAX_IMAGE_WIDTH = 600
MAX_IMAGE_HEIGHT = 400
DISPIMG_RE = re.compile(r"DISPIMG\s*\(\s*\"([^\"]+)\"", re.IGNORECASE)
INVALID_SHEET_CHARS_RE = re.compile(r"[\\/*?:\[\]]")

R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


@dataclass(frozen=True)
class DispimgImage:
    """图片数据及 WPS 保存的显示尺寸。尺寸单位为 EMU。"""

    data: bytes
    width_emu: int = 0
    height_emu: int = 0


@dataclass(frozen=True)
class NativeImageAnchor:
    col_from: int
    row_from: int
    col_to: int
    row_to: int
    col_off_from: int
    row_off_from: int
    col_off_to: int
    row_off_to: int


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _first_child(element: ET.Element | None, name: str) -> ET.Element | None:
    if element is None:
        return None
    return next((child for child in element.iter() if _local_name(child.tag) == name), None)


def extract_dispimg_id(value: object) -> str | None:
    """从单元格内容中提取 DISPIMG 的图片 ID。"""
    match = DISPIMG_RE.search(str(value or ""))
    return match.group(1) if match else None


def _safe_sheet_title(name: object, used: set[str]) -> str:
    title = INVALID_SHEET_CHARS_RE.sub("_", str(name or "Sheet"))[:31] or "Sheet"
    candidate = title
    suffix = 1
    while candidate in used:
        ending = f"_{suffix}"
        candidate = f"{title[:31 - len(ending)]}{ending}"
        suffix += 1
    used.add(candidate)
    return candidate


def _read_wps_package(data: bytes) -> dict[str, DispimgImage]:
    """读取 cellImages.xml 和其关系文件，返回 ID 到图片的映射。"""
    try:
        package = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile:
        return {}

    try:
        names = set(package.namelist())
        names_lower = {n.lower(): n for n in names}
        if "xl/cellimages.xml" not in names_lower:
            return {}

        relationships: dict[str, str] = {}
        rels_name = names_lower.get("xl/_rels/cellimages.xml.rels", "xl/_rels/cellImages.xml.rels")
        if rels_name in names:
            rels_root = ET.fromstring(package.read(rels_name))
            for relation in rels_root:
                rel_id = relation.get("Id")
                target = relation.get("Target")
                if rel_id and target:
                    relationships[rel_id] = target

        root = ET.fromstring(package.read(names_lower["xl/cellimages.xml"]))
        result: dict[str, DispimgImage] = {}
        for cell_image in root.iter():
            if _local_name(cell_image.tag) != "cellImage":
                continue
            picture = _first_child(cell_image, "pic")
            name_node = _first_child(picture, "cNvPr")
            blip = _first_child(picture, "blip")
            if name_node is None or blip is None:
                continue

            image_id = name_node.get("name", "")
            relation_id = blip.get(f"{{{R_NS}}}embed", "")
            target = relationships.get(relation_id)
            if not image_id or not target:
                continue

            # Relationship Target 相对于 xl/ 目录；同时兼容带 /xl/ 前缀的写法。
            target = target.replace("\\", "/")
            target_name = posixpath.basename(target)
            candidates = {
                posixpath.normpath(posixpath.join("xl", target)),
                posixpath.normpath(target.lstrip("/")),
                posixpath.join("xl", "media", target_name),
                posixpath.join("media", target_name),
            }
            image_name = next((candidate for candidate in candidates if candidate in names), None)
            if image_name is None:
                continue

            width_emu = height_emu = 0
            ext = _first_child(_first_child(picture, "xfrm"), "ext")
            if ext is not None:
                try:
                    width_emu = max(0, int(ext.get("cx", "0")))
                    height_emu = max(0, int(ext.get("cy", "0")))
                except ValueError:
                    pass
            result[image_id] = DispimgImage(package.read(image_name), width_emu, height_emu)
        return result
    except (ET.ParseError, KeyError, ValueError, RuntimeError):
        return {}
    finally:
        package.close()


def _read_xls_wps_package(path: Path) -> dict[str, DispimgImage]:
    """从旧式 XLS 的 OLE ETCellImageData 流读取 WPS 图片。"""
    try:
        import olefile
    except ImportError as exc:
        raise RuntimeError("处理 .xls 需要 olefile，请先执行: pip install olefile") from exc

    try:
        ole = olefile.OleFileIO(str(path))
    except (OSError, IOError) as exc:
        raise RuntimeError(f"无法打开 XLS 文件: {path.name}") from exc

    try:
        stream_path = next(
            (parts for parts in ole.listdir() if parts and parts[-1].lower() == "etcellimagedata"),
            None,
        )
        if stream_path is None:
            return {}
        return _read_wps_package(ole.openstream(stream_path).read())
    finally:
        ole.close()


def _read_xls_workbook_stream(path: Path) -> bytes:
    import olefile

    ole = olefile.OleFileIO(str(path))
    try:
        return ole.openstream("Workbook").read()
    finally:
        ole.close()


def _parse_native_image_anchors(path: Path) -> list[NativeImageAnchor]:
    """读取 XLS MsoDrawing 中的图片锚点，不处理文字框或图表。"""
    data = _read_xls_workbook_stream(path)
    object_types: dict[int, int] = {}
    index = 0
    while index + 4 <= len(data):
        opcode, record_length = struct.unpack("<HH", data[index:index + 4])
        if record_length > 100000:
            index += 2
            continue
        if opcode == 0x005D and record_length >= 6:
            object_types[index] = struct.unpack("<H", data[index + 8:index + 10])[0]
        index += 4 + record_length

    anchors: list[NativeImageAnchor] = []
    index = 0
    ordered_objects = sorted(object_types.items())
    while index + 4 <= len(data):
        opcode, record_length = struct.unpack("<HH", data[index:index + 4])
        if record_length > 100000:
            index += 2
            continue
        if opcode == 0x00EC:
            drawing = data[index + 4:index + 4 + record_length]
            next_types = [kind for offset, kind in ordered_objects if offset > index]
            if next_types and next_types[0] == 8:
                marker = drawing.find(b"\x10\xF0")
                if marker >= 0 and marker + 24 <= len(drawing):
                    marker_length = struct.unpack("<I", drawing[marker + 2:marker + 6])[0]
                    if marker_length == 18:
                        values = struct.unpack("<HHHHHHHHH", drawing[marker + 6:marker + 24])
                        _, col_from, col_off_from, row_from, row_off_from, col_to, col_off_to, row_to, row_off_to = values
                        anchors.append(NativeImageAnchor(
                            col_from, row_from, col_to, row_to,
                            col_off_from, row_off_from, col_off_to, row_off_to,
                        ))
        index += 4 + record_length
    return anchors


def _extract_native_images(path: Path) -> list[bytes]:
    """从 BIFF 的 MsoDrawing/Continue 记录重组并提取完整图片。"""
    data = _read_xls_workbook_stream(path)
    drawing_segments: list[bytearray] = []
    current: bytearray | None = None
    position = 0
    while position + 4 <= len(data):
        opcode, record_length = struct.unpack("<HH", data[position:position + 4])
        if record_length > 100000:
            position += 2
            continue
        payload = data[position + 4:position + 4 + record_length]
        if opcode in {0x00EB, 0x00EC}:
            current = bytearray(payload)
            drawing_segments.append(current)
        elif opcode == 0x003C and current is not None:
            current.extend(payload)
        else:
            current = None
        position += 4 + record_length

    images: list[tuple[int, bytes]] = []
    for segment in drawing_segments:
        drawing = bytes(segment)
        for marker, kind in ((b"\x89PNG\r\n\x1a\n", "png"), (b"\xff\xd8\xff", "jpeg")):
            search_from = 0
            while True:
                start = drawing.find(marker, search_from)
                if start < 0:
                    break
                end = -1
                if kind == "png":
                    offset = start + 8
                    while offset + 12 <= len(drawing):
                        chunk_length = struct.unpack(">I", drawing[offset:offset + 4])[0]
                        chunk_end = offset + 12 + chunk_length
                        if chunk_length > 50 * 1024 * 1024 or chunk_end > len(drawing):
                            break
                        if drawing[offset + 4:offset + 8] == b"IEND":
                            end = chunk_end
                            break
                        offset = chunk_end
                else:
                    end_marker = drawing.find(b"\xff\xd9", start + 3)
                    if end_marker >= 0:
                        end = end_marker + 2
                if end > start:
                    raw = drawing[start:end]
                    try:
                        # load() 会真正解码数据，避免把跨记录的残缺图片写入 XLSX。
                        from PIL import Image

                        with Image.open(io.BytesIO(raw)) as image:
                            image.load()
                        images.append((start, raw))
                    except Exception:
                        pass
                search_from = start + 1

    images.sort(key=lambda item: item[0])
    return [raw for _, raw in images]


def _prepare_image(data: bytes) -> tuple[bytes, int, int]:
    """确保图片格式是 openpyxl/Excel 通用的 PNG 或 JPEG。"""
    try:
        from PIL import Image

        with Image.open(io.BytesIO(data)) as image:
            width, height = image.size
            image.load()
            image_format = (image.format or "").upper()
            if image_format in {"BMP", "GIF", "JPEG", "PNG", "TIFF"}:
                return data, width, height

            converted = io.BytesIO()
            if image.mode not in {"RGB", "RGBA"}:
                image = image.convert("RGBA" if "A" in image.getbands() else "RGB")
            image.save(converted, format="PNG")
            return converted.getvalue(), width, height
    except Exception as exc:
        raise ValueError(f"无法识别图片格式: {exc}") from exc


def _column_width_pixels(ws, column: int) -> int:
    """将 1-based Excel 列宽近似换算为像素。"""
    from openpyxl.utils import get_column_letter

    dimension = ws.column_dimensions[get_column_letter(column)]
    if dimension.hidden:
        return 1
    width = dimension.width
    if width is None:
        width = ws.sheet_format.defaultColWidth or 13
    return max(1, int(width * 7 + 5))


def _row_height_pixels(ws, row: int) -> int:
    """将 1-based Excel 行高（磅）换算为像素。"""
    dimension = ws.row_dimensions[row]
    if dimension.hidden:
        return 1
    height = dimension.height or ws.sheet_format.defaultRowHeight or 15
    return max(1, int(height * 96 / 72))


def _cell_size_pixels(ws, row: int, column: int) -> tuple[int, int]:
    """返回目标单元格（支持合并单元格）的可用像素大小。"""
    first_row = last_row = row + 1
    first_column = last_column = column + 1
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row + 1 <= merged.max_row and merged.min_col <= column + 1 <= merged.max_col:
            first_row, last_row = merged.min_row, merged.max_row
            first_column, last_column = merged.min_col, merged.max_col
            break

    width = sum(_column_width_pixels(ws, current) for current in range(first_column, last_column + 1))
    height = sum(_row_height_pixels(ws, current) for current in range(first_row, last_row + 1))
    # 留出少量边距，避免图片压住单元格边框。
    return max(1, width - 4), max(1, height - 4)


def _image_display_size(info: DispimgImage, width: int, height: int, cell_size: tuple[int, int]) -> tuple[int, int]:
    if info.width_emu > 0 and info.height_emu > 0:
        source_width = info.width_emu / EMU_PER_PX
        source_height = info.height_emu / EMU_PER_PX
    else:
        ratio = min(MAX_IMAGE_WIDTH / max(width, 1), MAX_IMAGE_HEIGHT / max(height, 1), 1.0)
        source_width = width * ratio
        source_height = height * ratio

    cell_width, cell_height = cell_size
    ratio = min(cell_width / max(source_width, 1), cell_height / max(source_height, 1), 1.0)
    return max(1, int(source_width * ratio * EMU_PER_PX)), max(1, int(source_height * ratio * EMU_PER_PX))


def _add_floating_image(ws, info: DispimgImage, row: int, column: int) -> None:
    """将图片作为标准 oneCellAnchor 浮动对象添加到工作表。"""
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D

    image_data, width, height = _prepare_image(info.data)
    width_emu, height_emu = _image_display_size(info, width, height, _cell_size_pixels(ws, row, column))
    image = OpenpyxlImage(io.BytesIO(image_data))
    image.width = width_emu / EMU_PER_PX
    image.height = height_emu / EMU_PER_PX
    image.anchor = OneCellAnchor(
        _from=AnchorMarker(col=column, row=row, colOff=0, rowOff=0),
        ext=XDRPositiveSize2D(cx=width_emu, cy=height_emu),
    )
    ws.add_image(image)


def _add_native_floating_image(ws, raw_data: bytes, anchor: NativeImageAnchor) -> None:
    """按 XLS 原生锚点区域添加图片，保留其跨单元格的原始显示范围。"""
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D

    image_data, width, height = _prepare_image(raw_data)
    from_column = anchor.col_from + 1
    to_column = max(anchor.col_to + 1, from_column)
    from_row = anchor.row_from + 1
    to_row = max(anchor.row_to + 1, from_row)
    from_width = _column_width_pixels(ws, from_column)
    to_width = _column_width_pixels(ws, to_column)
    from_height = _row_height_pixels(ws, from_row)
    to_height = _row_height_pixels(ws, to_row)

    start_offset_x = from_width * min(anchor.col_off_from, 1023) / 1023
    end_offset_x = to_width * min(anchor.col_off_to, 1023) / 1023
    start_offset_y = from_height * min(anchor.row_off_from, 255) / 255
    end_offset_y = to_height * min(anchor.row_off_to, 255) / 255
    width_pixels = sum(_column_width_pixels(ws, column) for column in range(from_column, to_column))
    width_pixels += end_offset_x - start_offset_x
    height_pixels = sum(_row_height_pixels(ws, row) for row in range(from_row, to_row))
    height_pixels += end_offset_y - start_offset_y
    if width_pixels <= 1 or height_pixels <= 1:
        ratio = min(MAX_IMAGE_WIDTH / max(width, 1), MAX_IMAGE_HEIGHT / max(height, 1), 1.0)
        width_pixels, height_pixels = width * ratio, height * ratio

    width_emu = max(1, int(width_pixels * EMU_PER_PX))
    height_emu = max(1, int(height_pixels * EMU_PER_PX))
    image = OpenpyxlImage(io.BytesIO(image_data))
    image.width = width_pixels
    image.height = height_pixels
    image.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=anchor.col_from,
            row=anchor.row_from,
            colOff=int(start_offset_x * EMU_PER_PX),
            rowOff=int(start_offset_y * EMU_PER_PX),
        ),
        ext=XDRPositiveSize2D(cx=width_emu, cy=height_emu),
    )
    ws.add_image(image)


def _copy_xls_layout(src_ws, dst_ws) -> None:
    from openpyxl.utils import get_column_letter

    for rlo, rhi, clo, chi in src_ws.merged_cells:
        dst_ws.merge_cells(start_row=rlo + 1, start_column=clo + 1, end_row=rhi, end_column=chi)
    for index, info in getattr(src_ws, "colinfo_map", {}).items():
        if info.width:
            dst_ws.column_dimensions[get_column_letter(index + 1)].width = info.width / 256
    for index, info in getattr(src_ws, "rowinfo_map", {}).items():
        if info.height:
            dst_ws.row_dimensions[index + 1].height = info.height / 20


def _copy_xls_cell_value(cell, xl_wb):
    """复制值；日期转换为 datetime，错误单元格保持为空。"""
    import xlrd

    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR}:
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, xl_wb.datemode)
        except (ValueError, OverflowError):
            return cell.value
    return cell.value


def _find_dispimg_cells_xls(xl_wb) -> dict[int, list[tuple[int, int, str]]]:
    cells: dict[int, list[tuple[int, int, str]]] = {}
    for sheet_index in range(xl_wb.nsheets):
        sheet = xl_wb.sheet_by_index(sheet_index)
        for row in range(sheet.nrows):
            for column in range(sheet.ncols):
                image_id = extract_dispimg_id(sheet.cell_value(row, column))
                if image_id is None and hasattr(sheet, "cell_formula"):
                    try:
                        image_id = extract_dispimg_id(sheet.cell_formula(row, column))
                    except Exception:
                        pass
                if image_id:
                    cells.setdefault(sheet_index, []).append((row, column, image_id))
    return cells


def _convert_xls(path: Path, output_path: Path) -> tuple[int, int]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("处理 .xls 需要 xlrd，请先执行: pip install xlrd") from exc

    try:
        xl_wb = xlrd.open_workbook(str(path), formatting_info=True)
    except Exception as exc:
        raise RuntimeError(f"无法读取 XLS 文件: {exc}") from exc

    dispimg_cells = _find_dispimg_cells_xls(xl_wb)
    images = _read_xls_wps_package(path)
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    converted = 0
    missing = 0

    for sheet_index in range(xl_wb.nsheets):
        source = xl_wb.sheet_by_index(sheet_index)
        target = workbook.create_sheet(_safe_sheet_title(xl_wb.sheet_names()[sheet_index], used_titles))
        image_cells = {(row, column): image_id for row, column, image_id in dispimg_cells.get(sheet_index, [])}
        for row in range(source.nrows):
            for column in range(source.ncols):
                if (row, column) in image_cells:
                    continue
                value = _copy_xls_cell_value(source.cell(row, column), xl_wb)
                if value is not None:
                    target.cell(row=row + 1, column=column + 1, value=value)
        _copy_xls_layout(source, target)
        for row, column, image_id in dispimg_cells.get(sheet_index, []):
            info = images.get(image_id)
            if info is None:
                missing += 1
                continue
            try:
                _add_floating_image(target, info, row, column)
                converted += 1
            except ValueError:
                missing += 1

    # XLS 中还可能存在不使用 DISPIMG 的原生图片。它们本身不是公式，
    # 但也需要复制到标准 XLSX，否则会在转换后消失。
    native_anchors = _parse_native_image_anchors(path)
    native_images = _extract_native_images(path)
    native_count = min(len(native_anchors), len(native_images))
    if native_count:
        target_index = max(dispimg_cells, key=lambda index: len(dispimg_cells[index])) if dispimg_cells else 0
        target = workbook.worksheets[target_index]
        for anchor, raw_data in zip(native_anchors[:native_count], native_images[:native_count]):
            try:
                _add_native_floating_image(target, raw_data, anchor)
                converted += 1
            except ValueError:
                missing += 1
    _save_workbook(workbook, output_path)
    return converted, missing


def _read_xlsx_wps_images(path: Path) -> dict[str, DispimgImage]:
    """读取 XLSX 内的 cellImages.xml；普通 XLSX 没有该文件时返回空映射。"""
    try:
        with zipfile.ZipFile(path) as package:
            names_lower = {n.lower() for n in package.namelist()}
            if "xl/cellimages.xml" not in names_lower:
                return {}
            return _read_wps_package_from_xlsx(package)
    except (OSError, zipfile.BadZipFile):
        return {}


def _read_wps_package_from_xlsx(package: zipfile.ZipFile) -> dict[str, DispimgImage]:
    names = set(package.namelist())
    names_lower = {n.lower(): n for n in names}
    cell_xml = names_lower.get("xl/cellimages.xml", "")
    if not cell_xml:
        return {}
    # 将 XLSX 包内的相关文件复制为最小 ZIP，再复用同一解析器。
    rels_xml = names_lower.get("xl/_rels/cellimages.xml.rels", "xl/_rels/cellImages.xml.rels")
    memory = io.BytesIO()
    with zipfile.ZipFile(memory, "w", zipfile.ZIP_DEFLATED) as output:
        for name in (cell_xml, rels_xml):
            if name in names:
                output.writestr(name, package.read(name))
        for name in names:
            if name.startswith("xl/media/"):
                output.writestr(name, package.read(name))
    return _read_wps_package(memory.getvalue())


def _save_workbook(workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(prefix=".convert_", suffix=".xlsx", dir=output_path.parent, delete=False) as handle:
        temporary_path = Path(handle.name)
    try:
        workbook.save(str(temporary_path))
        temporary_path.replace(output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def convert_file(path: Path, output_dir: Path) -> tuple[int, int, Path]:
    output_path = output_dir / f"{path.stem}.xlsx"
    if output_path.resolve() == path.resolve():
        raise ValueError("输出文件不能覆盖输入文件，请使用其他 --output-dir")
    if path.suffix.lower() == ".xls":
        converted, missing = _convert_xls(path, output_path)
    elif path.suffix.lower() == ".xlsx":
        # 读取 WPS 自定义图片包，再交给 openpyxl 保留标准 XLSX 内容。
        images = _read_xlsx_wps_images(path)
        converted, missing = _convert_xlsx_with_images(path, output_path, images)
    else:
        raise ValueError(f"不支持的文件类型: {path.suffix}")
    return converted, missing, output_path


def _convert_xlsx_with_images(path: Path, output_path: Path, images: dict[str, DispimgImage]) -> tuple[int, int]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), data_only=False)
    converted = 0
    missing = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                image_id = extract_dispimg_id(cell.value)
                if not image_id:
                    continue
                cell.value = None
                info = images.get(image_id)
                if info is None:
                    missing += 1
                    continue
                try:
                    _add_floating_image(worksheet, info, cell.row - 1, cell.column - 1)
                    converted += 1
                except ValueError:
                    missing += 1
    _save_workbook(workbook, output_path)
    return converted, missing


def _iter_input_files(input_path: Path, output_dir: Path, recursive: bool) -> Iterable[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() in {".xls", ".xlsx"} and not input_path.name.startswith("~$"):
            yield input_path
        return
    pattern = "**/*" if recursive else "*"
    output_dir = output_dir.resolve()
    for path in sorted(input_path.glob(pattern)):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in {".xls", ".xlsx"}:
            continue
        try:
            path.resolve().relative_to(output_dir)
        except ValueError:
            yield path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="将 WPS DISPIMG 图片转换为标准 XLSX 浮动图片")
    parser.add_argument("input", nargs="?", type=Path, default=Path(__file__).parent, help="输入 XLS/XLSX 文件或目录")
    parser.add_argument("-o", "--output-dir", type=Path, help="输出目录，默认是输入目录下的 finishexcel")
    parser.add_argument("-r", "--recursive", action="store_true", help="输入为目录时递归处理子目录")
    args = parser.parse_args(argv)

    input_path = args.input.resolve()
    if not input_path.exists():
        parser.error(f"输入路径不存在: {input_path}")
    output_dir = (args.output_dir or (input_path.parent if input_path.is_file() else input_path) / "finishexcel").resolve()
    files = list(_iter_input_files(input_path, output_dir, args.recursive))
    if not files:
        print("未找到 .xls 或 .xlsx 文件。")
        return 0

    print(f"输入: {input_path}")
    print(f"输出: {output_dir}")
    success = failed = converted_total = missing_total = 0
    for path in files:
        try:
            converted, missing, output_path = convert_file(path, output_dir)
            print(f"[完成] {path.name} -> {output_path.name}，转换 {converted} 张" + (f"，缺失 {missing} 张" if missing else ""))
            success += 1
            converted_total += converted
            missing_total += missing
        except Exception as exc:
            failed += 1
            print(f"[失败] {path.name}: {exc}", file=sys.stderr)

    print(f"处理完成：成功 {success}，失败 {failed}，转换 {converted_total} 张，缺失 {missing_total} 张")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
