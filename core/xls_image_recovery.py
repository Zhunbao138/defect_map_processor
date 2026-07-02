"""xls 图片恢复模块。

从 xls 文件中提取所有嵌入图片，并尝试通过 OCR 反向匹配钢板号，
建立图片与记录的关联。然后将图片嵌入到新生成的 xlsx 中。

策略:
1. 扫描 xls 文件二进制流，提取所有 JPEG/PNG 图片
2. 对每张图做 OCR 识别（在图片的文字区域找钢板号）
3. 根据钢板号匹配 sheet2 的某一行
4. 用 openpyxl 加载 xlsx，把图片插入到 F 列对应行
"""
from __future__ import annotations

import io
import re
import struct
from pathlib import Path
from typing import Any

import olefile
from PIL import Image


def extract_images_from_xls_binary(xls_path: str | Path) -> list[bytes]:
    """直接从 xls 二进制扫描并提取所有 JPEG/PNG 图片。

    Returns:
        图片二进制列表, 每项是完整的 JPEG 或 PNG 数据
    """
    with open(xls_path, "rb") as f:
        data = f.read()

    images = []

    # 找所有 JPEG
    pos = 0
    while True:
        idx = data.find(b"\xff\xd8\xff", pos)
        if idx == -1:
            break
        end = data.find(b"\xff\xd9", idx + 3)
        if end == -1 or end - idx > 50 * 1024 * 1024:
            pos = idx + 1
            continue
        jpg = data[idx : end + 2]
        # 验证
        try:
            img = Image.open(io.BytesIO(jpg))
            img.verify()
            img2 = Image.open(io.BytesIO(jpg))
            if img2.size[0] >= 50 and img2.size[1] >= 50:
                images.append(jpg)
        except Exception:
            pass
        pos = idx + 1

    return images


def extract_plate_no_from_image(img_bytes: bytes) -> str | None:
    """从图片中 OCR 识别钢板号。

    钢板号格式: 14 位数字 (例如 26102421340201)
    """
    try:
        import easyocr
        import numpy as np

        reader = easyocr.Reader(["en"], gpu=False, verbose=False)
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")

        # 缩小图片 (大图OCR太慢)
        w, h = img.size
        max_size = 1500
        if w > max_size or h > max_size:
            ratio = max_size / max(w, h)
            img = img.resize((int(w * ratio), int(h * ratio)))

        arr = np.array(img)
        result = reader.readtext(arr, detail=0)
        text = " ".join(result)

        # 找 14 位连续数字
        m = re.search(r"\b(\d{14})\b", text.replace(" ", ""))
        if m:
            return m.group(1)

        # 找带分隔符的
        m = re.search(r"(\d[\dxX]{12,16})", text.replace(" ", ""))
        if m:
            return re.sub(r"[^\d]", "", m.group(1))
    except Exception:
        pass
    return None


def _pil_to_numpy(img):
    import numpy as np

    return np.array(img)


def match_images_to_records(
    images: list[bytes],
    records: list[dict],
) -> dict[int, list[bytes]]:
    """通过 OCR 反向匹配图片到记录。

    Returns:
        {row_index: [img_bytes, ...]}
    """
    matched: dict[int, list[bytes]] = {}
    plate_to_row = {
        r.get("钢板号", "").strip(): r.get("row_index")
        for r in records
        if r.get("钢板号")
    }

    for img in images:
        plate_no = extract_plate_no_from_image(img)
        if plate_no and plate_no in plate_to_row:
            row = plate_to_row[plate_no]
            matched.setdefault(row, []).append(img)

    return matched


def embed_images_into_xlsx(
    xlsx_path: str | Path,
    images_by_row: dict[int, list[bytes]],
    images_dir: str | Path,
    column: int = 6,  # F 列
) -> list[dict]:
    """把图片嵌入到 xlsx 的指定列。

    Args:
        xlsx_path: xlsx 文件路径
        images_by_row: {row_index: [img_bytes, ...]}
        images_dir: 图片保存目录
        column: 列号 (1-based, 默认 6 = F)

    Returns:
        图片元数据列表
    """
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage

    images_dir = Path(images_dir)
    images_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path)
    # 总是选第二个表 (按索引), 不再按名字匹配
    if len(wb.sheetnames) < 2:
        raise ValueError(
            f"xlsx 至少需要 2 个 sheet (当前 {len(wb.sheetnames)} 个: {wb.sheetnames})"
        )
    sheet = wb[wb.sheetnames[1]]

    results = []
    for row_idx, imgs in images_by_row.items():
        for img_num, img_bytes in enumerate(imgs[:2], 1):
            # 保存为文件
            img_path = images_dir / f"row{row_idx}_img{img_num}.png"
            try:
                pil_img = Image.open(io.BytesIO(img_bytes))
                pil_img.save(str(img_path), "PNG")
            except Exception:
                # 直接保存原始字节
                ext = "jpg" if img_bytes[:3] == b"\xff\xd8\xff" else "png"
                img_path = images_dir / f"row{row_idx}_img{img_num}.{ext}"
                img_path.write_bytes(img_bytes)

            # 嵌入 xlsx
            cell = sheet.cell(row=row_idx, column=column)
            oimg = OpenpyxlImage(str(img_path))
            oimg.width = 200
            oimg.height = 150
            sheet.add_image(oimg, cell.coordinate)

            results.append(
                {
                    "row_index": row_idx,
                    "image_index": img_num,
                    "file_path": str(img_path),
                    "format": "png",
                    "width": pil_img.size[0] if "pil_img" in dir() else 0,
                    "height": pil_img.size[1] if "pil_img" in dir() else 0,
                }
            )

    wb.save(xlsx_path)
    return results


if __name__ == "__main__":
    xls_path = "/home/qing/文档/t1/2026-6-11国家项目数据采集日常工作搜集表(中板厂).xls"
    images = extract_images_from_xls_binary(xls_path)
    print(f"提取了 {len(images)} 张图")
    for i, img in enumerate(images):
        pil = Image.open(io.BytesIO(img))
        print(f"  图{i+1}: {pil.size} ({len(img)/1024:.1f} KB)")