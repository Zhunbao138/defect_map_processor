"""xls/xlsx 文件导入与数据读取模块。

支持:
- 自动识别 .xls / .xlsx 格式
- .xls 自动转换为 .xlsx
- 读取指定 sheet (默认 sheet2)
- 解析缺陷记录 (序号, 钢板号, 钢种, 类别, 缺陷分析等)
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.utils import get_column_letter


def parse_notes_to_params(notes: str) -> dict[str, str]:
    """从附加说明文本解析 OCR 参数。

    支持的格式:
        左上角材料尺寸：12000×2430×30
        缺陷中心X[mm]:8412
        缺陷中心Y[mm]:1839
        缺陷长度[mm]:1669.4
        缺陷宽度[mm]:162.6
        缺陷深度[mm]:14.3
    """
    if not notes:
        return {}

    params = {}

    # 材料尺寸: 多种分隔符
    m = re.search(r"材料尺寸[：:]\s*(\d+)\s*[×xX\*·]\s*(\d+)\s*[×xX\*·]\s*(\d+)", notes)
    if m:
        params["材料尺寸"] = f"{m.group(1)}×{m.group(2)}×{m.group(3)}"

    # 缺陷中心X
    m = re.search(r"缺陷中心\s*X\s*\[mm\]\s*[:：]\s*([\d.]+)", notes)
    if m:
        params["缺陷中心X"] = m.group(1)

    # 缺陷中心Y
    m = re.search(r"缺陷中心\s*Y\s*\[mm\]\s*[:：]\s*([\d.]+)", notes)
    if m:
        params["缺陷中心Y"] = m.group(1)

    # 缺陷长度
    m = re.search(r"缺陷长度\s*\[mm\]\s*[:：]\s*([\d.]+)", notes)
    if m:
        params["缺陷长度"] = m.group(1)

    # 缺陷宽度
    m = re.search(r"缺陷宽度\s*\[mm\]\s*[:：]\s*([\d.]+)", notes)
    if m:
        params["缺陷宽度"] = m.group(1)

    # 缺陷深度
    m = re.search(r"缺陷深度\s*\[mm\]\s*[:：]\s*([\d.]+)", notes)
    if m:
        params["缺陷深度"] = m.group(1)

    return params


class ExcelImporter:
    """Excel 文件导入器。

    处理 .xls / .xlsx 两种格式，提取 sheet2 中的缺陷记录数据。
    """

    # sheet2 表头映射 (xlsx 只有 8 列，没有"附加说明"列)
    # 缺陷数据（材料尺寸、缺陷中心、长度宽度深度等）通过 OCR 识别图片得到
    COLUMN_MAPPING = {
        0: "序号",
        1: "生产厂",
        2: "钢板号",
        3: "钢种",
        4: "类别",
        5: "缺陷图谱",   # F 列，2 张图
        6: "缺陷照片",   # G 列
        7: "缺陷分析",
    }

    # 数据起始行: 跳过表头(行1)、表头字段名(行2)、负责单位(行3)、要求(行4)、示意行(行5-6)，从第7行开始
    DATA_START_ROW = 7

    # 需要排除的"非数据"标记（出现在"序号"或"生产厂"列的关键词）
    EXCLUDED_KEYWORDS = ("示意", "示例", "example", "占位")

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        self.suffix = self.file_path.suffix.lower()
        self.workbook_xls: xlrd.Book | None = None
        self.workbook_xlsx: OpenpyxlWorkbook | None = None
        self.converted_xlsx_path: Path | None = None
        self.extracted_images: list[bytes] = []  # 从 xls 提取的图片二进制

    def import_file(self) -> Path:
        """导入 xlsx 文件。

        只支持 .xlsx 格式。如需处理 .xls, 请先用 WPS/Excel 另存为 .xlsx。
        """
        if self.suffix == ".xls":
            raise ValueError(
                "不支持 .xls 格式! 请先用 WPS/Excel 打开并另存为 .xlsx 后再上传。"
            )
        if self.suffix == ".xlsx":
            return self.file_path
        raise ValueError(f"不支持的文件格式: {self.suffix}，仅支持 .xlsx")

    def _convert_xls_to_xlsx(self) -> Path:
        """已废弃: 不再支持 xls 格式转换。"""
        raise ValueError(
            "不支持 .xls 格式! 请先用 WPS/Excel 打开并另存为 .xlsx 后再上传。"
        )

    def _embed_images_to_xlsx(self, xlsx_path: Path):
        """(已废弃) 此方法不再使用。"""
        pass

    def _assign_images_to_records(self, xlsx_path: Path):
        """(已废弃) 此方法不再使用。"""
        pass

    def read_sheet2_records(self) -> list[dict[str, Any]]:
        """读取 sheet2 的所有缺陷记录。

        自动过滤:
        - 表头/负责单位/要求等说明行（已通过 DATA_START_ROW 跳过）
        - 序号列或生产厂列含"示意/示例/占位"等关键词的行
        - 无钢板号或无序号的空行

        返回格式:
        [
            {
                "row_index": 7,            # 原表行号（1-based）
                "序号": "1",
                "生产厂": "中板厂",
                "钢板号": "26102421340201",
                "钢种": "JX22145-2025 X60MS",
                "类别": "管线",
                "缺陷分析": "显孔",
                "附加说明": "...",
            },
            ...
        ]
        """
        def is_excluded(rec: dict) -> bool:
            """检查该行是否为示意/示例行。"""
            for col in ("序号", "生产厂"):
                v = rec.get(col, "")
                for kw in self.EXCLUDED_KEYWORDS:
                    if kw in v:
                        return True
            # 序号必须是数字（1, 2, 3...）
            seq = rec.get("序号", "").strip()
            if not seq or not seq.split(".")[0].isdigit():
                return True
            return False

        # 只支持 xlsx
        self.import_file()
        from openpyxl import load_workbook as load_xlsx
        wb = load_xlsx(str(self.converted_xlsx_path or self.file_path), data_only=True)
        ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb[wb.sheetnames[1]]
        records = []
        for row_idx in range(self.DATA_START_ROW, ws.max_row + 1):
            record = self._parse_xlsx_row(ws, row_idx)
            if not record or not record.get("钢板号"):
                continue
            if is_excluded(record):
                continue
            records.append(record)
        return records

    def _parse_xls_row(self, sheet: xlrd.sheet.Sheet, row_idx: int) -> dict[str, Any]:
        """已废弃: 不再支持 xls 格式。"""
        raise NotImplementedError("不支持 .xls 格式")
        record: dict[str, Any] = {"row_index": row_idx + 1}
        for col_idx, col_name in self.COLUMN_MAPPING.items():
            try:
                val = sheet.cell(row_idx, col_idx).value
                if val is not None and str(val).strip():
                    record[col_name] = str(val).strip()
                else:
                    record[col_name] = ""
            except Exception:
                record[col_name] = ""
        # 缺陷数据由 OCR 识别图片得到，不从表格列读
        record["缺陷数据"] = {}
        record["ocr_raw_text"] = ""
        return record

    def _parse_xlsx_row(self, ws, row_idx: int) -> dict[str, Any]:
        """从 openpyxl ws 解析单行。"""
        record: dict[str, Any] = {"row_index": row_idx}
        for col_idx, col_name in self.COLUMN_MAPPING.items():
            try:
                val = ws.cell(row=row_idx, column=col_idx + 1).value
                if val is not None and str(val).strip():
                    record[col_name] = str(val).strip()
                else:
                    record[col_name] = ""
            except Exception:
                record[col_name] = ""
        # 缺陷数据由 OCR 识别图片得到，不从表格列读
        record["缺陷数据"] = {}
        record["ocr_raw_text"] = ""
        return record


def convert_to_xlsx(xls_path: str | Path, output_dir: str | Path | None = None) -> Path:
    """便捷函数：xls 转 xlsx。"""
    importer = ExcelImporter(xls_path)
    return importer.import_file()


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("用法: python importer.py <xls/xlsx文件路径>")
        sys.exit(1)

    importer = ExcelImporter(sys.argv[1])
    xlsx_path = importer.import_file()
    print(f"已转换为: {xlsx_path}")

    records = importer.read_sheet2_records()
    print(f"\n读取到 {len(records)} 条缺陷记录\n")
    for r in records[:5]:
        print(f"  行{r['row_index']}: {r.get('钢板号', '')} | {r.get('钢种', '')} | {r.get('类别', '')}")
    print(f"  ...")
    for r in records[-3:]:
        print(f"  行{r['row_index']}: {r.get('钢板号', '')} | {r.get('钢种', '')} | {r.get('类别', '')}")