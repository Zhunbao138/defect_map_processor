"""图片提取模块。

支持从 xls / xlsx 中提取嵌入的图片：

1. xlsx 格式: 通过 openpyxl + zipfile 直接读取 xl/media 目录下的嵌入图片
2. xls 格式: 通过 olefile + BIFF 流扫描，提取 JPEG / PNG / EMF 等格式

每张图片保存为单独文件，文件命名规则:
    row_{row_idx}_img_{image_idx}.{ext}

返回结构:
    [
        {
            "row_index": int,         # 在 sheet2 中对应的行号
            "image_index": int,       # 该行的第几张图 (从1开始)
            "file_path": str,         # 保存的本地路径
            "format": str,            # 原格式 jpeg/png/emf
        },
        ...
    ]
"""
from __future__ import annotations

import io
import os
import re
import shutil
import struct
import zipfile
from pathlib import Path

import olefile
from PIL import Image
from openpyxl_image_loader import SheetImageLoader
from openpyxl import load_workbook

# 让用户友好提示缺包
try:
    from openpyxl_image_loader import SheetImageLoader
except ImportError:
    SheetImageLoader = None


class ImageExtractor:
    """从 Excel 文件提取嵌入图片。"""

    # 常见图片格式头
    IMAGE_MARKERS = [
        # (name, header_bytes, find_end_method)
        ("jpeg", b"\xff\xd8\xff", "find_jpeg_end"),
        ("png", b"\x89PNG\r\n\x1a\n", "find_png_end"),
        ("gif", b"GIF8", "find_gif_end"),
        ("bmp", b"BM", "parse_bmp_end"),
        ("tiff_le", b"II*\x00", "find_tiff_end"),
        ("tiff_be", b"MM\x00*", "find_tiff_end"),
        # EMF: 0x01 0x00 0x00 0x00 0x58 = signature record (28 bytes header)
        ("emf", b"\x01\x00\x00\x00\x58", "parse_emf_end"),
    ]

    # 单张图片大小上限 (50MB)
    MAX_IMAGE_SIZE = 50 * 1024 * 1024

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        self.suffix = self.file_path.suffix.lower()

    def extract(
        self, output_dir: str | Path, sheet_name: str = "Sheet2"
    ) -> list[dict]:
        """提取所有图片到 output_dir。

        返回每张图片的元数据列表。
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        if self.suffix == ".xlsx":
            return self._extract_xlsx(output_dir, sheet_name)
        elif self.suffix == ".xls":
            return self._extract_xls(output_dir)
        else:
            raise ValueError(f"不支持的文件格式: {self.suffix}")

    # ------------------------------------------------------------------ #
    # xlsx 提取 (通过 openpyxl_image_loader)
    # ------------------------------------------------------------------ #
    def _extract_xlsx(self, output_dir: Path, sheet_name: str) -> list[dict]:
        """从 xlsx 提取图片。

        原理：通过 openpyxl 的 _images 列表获取所有图片的真实锚点 (行/列)。
        openpyxl-image-loader 在某些情况下会丢失图片, 所以我们直接用 sheet._images。
        """
        wb = load_workbook(str(self.file_path))
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"找不到 sheet: {sheet_name}，可选: {wb.sheetnames}")
        ws = wb[sheet_name]

        results = []
        # 只提取 F 列 (col=5) 的图片, 即"缺陷图谱"列
        # F 列本身就有两张图 (整体图谱 + 局部放大图谱)
        TARGET_COL = 5  # 0-based, F 列

        # 收集每行的所有图片
        row_to_images: dict[int, list] = {}
        for img_idx, img_obj in enumerate(ws._images):
            anchor = img_obj.anchor
            if not hasattr(anchor, "_from"):
                continue
            from_anchor = anchor._from
            row_idx = from_anchor.row + 1  # 0-based → 1-based
            col_idx = from_anchor.col

            # 只看 F 列
            if col_idx != TARGET_COL:
                continue

            row_to_images.setdefault(row_idx, []).append(img_obj)

        # 每张图归属于其锚点行 (WPS 中直接显示的位置)
        for row_idx in sorted(row_to_images.keys()):
            imgs_in_row = row_to_images[row_idx]
            for img_seq, img_obj in enumerate(imgs_in_row, 1):
                col_name = "缺陷图谱"

                # 保存图片
                img_data = img_obj._data() if hasattr(img_obj, '_data') else None
                if img_data is None:
                    continue

                from PIL import Image as PILImage
                import io

                # 保留原始字节，不重新编码（避免任何重压缩或色板变化）
                # 根据原始数据判断真实扩展名
                if img_data[:8] == b"\x89PNG\r\n\x1a\n":
                    ext = "png"
                elif img_data[:2] == b"\xff\xd8":
                    ext = "jpg"
                elif img_data[:4] == b"GIF8":
                    ext = "gif"
                elif img_data[:2] == b"BM":
                    ext = "bmp"
                else:
                    ext = "png"  # 默认

                file_path = (
                    output_dir
                    / f"row{row_idx:03d}_{col_name}_{img_seq}.{ext}"
                )
                file_path.write_bytes(img_data)

                # 用 PIL 读尺寸（不写盘）
                try:
                    pil_img = PILImage.open(io.BytesIO(img_data))
                    w, h = pil_img.size
                except Exception:
                    w, h = 0, 0

                results.append(
                    {
                        "row_index": row_idx,
                        "column": col_name,
                        "image_index": img_seq,
                        "file_path": str(file_path),
                        "format": "png",
                        "width": w,
                        "height": h,
                    }
                )

        # 按 row_index 和 image_index 排序
        results.sort(key=lambda x: (x["row_index"], x["column"], x["image_index"]))
        return results

    # ------------------------------------------------------------------ #
    # xls 提取 (直接扫描 BIFF 流)
    # ------------------------------------------------------------------ #
    def _extract_xls(self, output_dir: Path) -> list[dict]:
        """从 xls 提取嵌入图片。

        BIFF8 格式的图片存储说明：
        - 文件是 OLE2 复合文档
        - Workbook stream 中包含 PICTURE (0x00E0) 记录作为图片元数据
        - 实际的二进制图片数据通常通过:
          a) 紧跟在 OBJ (0x005D) 记录后的子记录链
          b) 直接以 JPEG/PNG 等标记嵌入在 stream 中
        - 这里采用方法 b：扫描整个 stream 找所有合法的图片块
        """
        # 优先尝试方法 a：解析 PICTURE 记录链
        results = self._extract_xls_picture_records(output_dir)
        if results:
            return results

        # 回退到方法 b：扫描二进制流
        return self._extract_xls_scan_stream(output_dir)

    def _extract_xls_picture_records(self, output_dir: Path) -> list[dict]:
        """解析 PICTURE 记录，尝试组合成完整图片。

        BIFF 中图片对象结构:
            OBJ (0x005D, ftCmo, ftPict, ftEnd)
            -> 后面跟 IMDATA 子记录或外部存储
        """
        if not olefile.isOleFile(str(self.file_path)):
            return []

        ole = olefile.OleFileIO(str(self.file_path))
        try:
            stream = ole.openstream("Workbook").read()
        except Exception:
            ole.close()
            return []
        ole.close()

        results = []
        # 扫描找每个 OBJ 记录 + 紧接的可疑图片数据
        i = 0
        obj_index = 0
        while i < len(stream) - 4:
            opcode = struct.unpack("<H", stream[i : i + 2])[0]
            rec_len = struct.unpack("<H", stream[i + 2 : i + 4])[0]
            if rec_len > 100000000 or rec_len < 0:
                i += 2
                continue

            if opcode == 0x005D:  # OBJ
                obj_index += 1
                # OBJ 后 4 字节是 cmo, 然后 sub-records
                # ft (2字节) + len (2字节) + data
                sub_i = i + 4 + 4  # skip cmo
                while sub_i < i + 4 + rec_len - 4:
                    if sub_i + 4 > len(stream):
                        break
                    ft = struct.unpack("<H", stream[sub_i : sub_i + 2])[0]
                    ft_len = struct.unpack("<H", stream[sub_i + 2 : sub_i + 4])[0]
                    if ft_len > 100000000 or ft_len < 0:
                        break
                    if ft == 0x0021:  # ftEnd
                        break
                    # ftPictFmla (0x0009) 或 ftPict (0x000A)
                    if ft in (0x0009, 0x000A):
                        # 后面通常有图片二进制
                        # CONTINUE record (0x003C) 紧接其后的才是图片数据
                        pass
                    sub_i += 4 + ft_len

            i += 4 + rec_len

        # 上面只是分析了 OBJ 结构; 实际图片数据用 scan_stream 更可靠
        return []

    def _extract_xls_scan_stream(self, output_dir: Path) -> list[dict]:
        """扫描整个 xls 二进制找所有合法图片。"""
        with open(self.file_path, "rb") as f:
            data = f.read()

        results = []
        seen_positions = set()  # 防止重叠

        # 收集所有图片起始位置
        candidates = []
        for fmt_name, marker, _ in self.IMAGE_MARKERS:
            pos = 0
            while True:
                idx = data.find(marker, pos)
                if idx == -1:
                    break
                candidates.append((idx, fmt_name, marker))
                pos = idx + 1

        # 按位置排序
        candidates.sort()

        # 对每个候选位置尝试提取完整图片
        for idx, fmt_name, marker in candidates:
            # 跳过重叠
            if any(abs(idx - p) < 50 for p in seen_positions):
                continue

            img_data = self._extract_image_at(data, idx, fmt_name)
            if img_data is None:
                continue

            try:
                img = Image.open(io.BytesIO(img_data))
                img.verify()
                img = Image.open(io.BytesIO(img_data))
                width, height = img.size
                if width < 50 or height < 50:
                    continue
            except Exception:
                continue

            # 保存
            file_path = output_dir / f"xls_image_{len(results):03d}.{fmt_name}"
            with open(file_path, "wb") as f:
                f.write(img_data)

            seen_positions.add(idx)
            results.append(
                {
                    "row_index": None,  # xls 中难以关联到具体行
                    "column": "缺陷图谱",
                    "image_index": len(results) + 1,
                    "file_path": str(file_path),
                    "format": fmt_name,
                    "width": width,
                    "height": height,
                }
            )

        return results

    def _extract_image_at(self, data: bytes, pos: int, fmt_name: str) -> bytes | None:
        """根据格式从 pos 开始提取完整图片数据。"""
        try:
            if fmt_name == "jpeg":
                end = data.find(b"\xff\xd9", pos + 3)
                if end == -1 or end - pos > self.MAX_IMAGE_SIZE:
                    return None
                return data[pos : end + 2]
            elif fmt_name == "png":
                end = data.find(b"IEND", pos)
                if end == -1 or end - pos > self.MAX_IMAGE_SIZE:
                    return None
                return data[pos : end + 8]  # IEND + 4字节CRC
            elif fmt_name == "gif":
                # GIF 结束标记: trailer byte 0x3B
                end = data.find(b"\x3b", pos + 10)
                if end == -1 or end - pos > self.MAX_IMAGE_SIZE:
                    return None
                return data[pos : end + 1]
            elif fmt_name == "bmp":
                # BMP 文件大小在 offset 2-6 (4 字节 LE)
                if pos + 6 > len(data):
                    return None
                file_size = struct.unpack("<I", data[pos + 2 : pos + 6])[0]
                if file_size < 100 or file_size > self.MAX_IMAGE_SIZE:
                    return None
                return data[pos : pos + file_size]
            elif fmt_name in ("tiff_le", "tiff_be"):
                # TIFF 没有明确的 end marker, 取到下一个图片或 N 张图
                # 简化: 取到下一个 \xff\xd8 或 \x89PNG 之前
                end = len(data)
                for stop_marker in [b"\xff\xd8\xff", b"\x89PNG", b"GIF8"]:
                    idx = data.find(stop_marker, pos + 100)
                    if idx != -1 and idx < end:
                        end = idx
                if end - pos > self.MAX_IMAGE_SIZE:
                    return None
                return data[pos:end]
            elif fmt_name == "emf":
                # EMF: header 28 bytes, then records until EOF
                # Get file size from header
                if pos + 48 > len(data):
                    return None
                file_size = struct.unpack("<I", data[pos + 40 : pos + 44])[0]
                if file_size < 100 or file_size > self.MAX_IMAGE_SIZE:
                    return None
                # Convert EMF to PNG via libreoffice or PIL
                emf_data = data[pos : pos + file_size]
                return self._convert_emf_to_png(emf_data)
        except Exception:
            return None
        return None

    def _convert_emf_to_png(self, emf_data: bytes) -> bytes | None:
        """将 EMF 转换为 PNG。

        优先使用 LibreOffice (soffice headless)，回退到 PIL。
        PIL 实际上不支持 EMF，所以这里先用临时文件 + soffice。
        """
        import tempfile

        try:
            with tempfile.NamedTemporaryFile(suffix=".emf", delete=False) as tmp_in:
                tmp_in.write(emf_data)
                tmp_in_path = tmp_in.name

            # 尝试用 soffice 转换
            tmp_out_dir = tempfile.mkdtemp()
            subprocess.run(
                [
                    "soffice",
                    "--headless",
                    "--convert-to",
                    "png",
                    "--outdir",
                    tmp_out_dir,
                    tmp_in_path,
                ],
                check=True,
                capture_output=True,
                timeout=30,
            )
            png_file = Path(tmp_out_dir) / (Path(tmp_in_path).stem + ".png")
            if png_file.exists():
                png_data = png_file.read_bytes()
                os.remove(tmp_in_path)
                shutil.rmtree(tmp_out_dir)
                return png_data
        except Exception:
            pass
        return None


# ============================================================================
# 辅助函数
# ============================================================================
def extract_images_from_excel(
    file_path: str | Path,
    output_dir: str | Path,
    sheet_name: str = "Sheet2",
) -> list[dict]:
    """便捷函数：从 Excel 文件提取图片。"""
    extractor = ImageExtractor(file_path)
    return extractor.extract(output_dir, sheet_name)


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python image_extractor.py <xls/xlsx文件路径> [输出目录]")
        sys.exit(1)

    file_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "extracted_images"
    results = extract_images_from_excel(file_path, output_dir)
    print(f"\n共提取 {len(results)} 张图片:")
    for r in results[:10]:
        print(f"  行{r['row_index']} {r['column']}-{r['image_index']}: {r['file_path']} ({r['width']}×{r['height']})")
    if len(results) > 10:
        print(f"  ... 共 {len(results)} 张")