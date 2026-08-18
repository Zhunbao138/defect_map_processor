"""cscan 文档处理模块 (中厚板卷厂).

- 提取 F/G 列原图 → 切出 3 个子图 (table / ascan / cscan) → 保存
- OCR 13 列缺陷表格 → 13 个字段每行
- OCR 板信息 panel → 板号/钢种/厚度/长度/宽度 等字段

复用 core/view_splitter.detect_and_crop_cscan_views 的算法.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ============================================================
# 13 列缺陷表格列定义
# ============================================================
DEFECT_TABLE_COLS = (
    "序号", "X起始", "X终止", "X中点", "X长度",
    "Y起始", "Y终止", "Y中点", "Y长度",
    "面积", "类型", "深度", "幅值",
)


# ============================================================
# 单张原图切图 (包装 view_splitter.detect_and_crop_cscan_views)
# ============================================================
def split_cscan_views(
    image_path: str | Path,
    output_dir: str | Path,
    prefix: str = "",
) -> dict[str, Path]:
    """切出 cscan 原图里的 4 个子图, 加 prefix 前缀 (F/G) 保存.

    Args:
        image_path: 原图路径
        output_dir: 子图保存目录
        prefix: 文件名前缀 (如 "F" / "G", 留空则不加)

    Returns:
        {"table": Path, "ascan": Path, "cscan": Path, "warnings": [...]}
    """
    from .view_splitter import detect_and_crop_cscan_views

    image_path = Path(image_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    result = detect_and_crop_cscan_views(image_path, output_dir, draw_annotations=False)
    out: dict[str, Any] = {"warnings": result.get("warnings", [])}
    for en_name, src_path in result.get("views", {}).items():
        out[en_name] = Path(src_path)
    return out


# ============================================================
# 提取 F/G 列所有原图并切图
# ============================================================
def _extract_date_from_title(ws) -> str | None:
    """从 sheet 标题 row 1 提取日期标签 (如 '4.21')."""
    title = str(ws.cell(1, 1).value or "")
    m = re.search(r'（(\d{4})-(\d{1,2})-(\d{1,2})）', title)
    return f"{m.group(2)}.{int(m.group(3))}" if m else None


def find_cscan_sheets(xlsx_path: str | Path) -> list[tuple[str, str]]:
    """模板二: 找出日期格式 sheet, 找不到则用模板三方法兜底.

    返回 [(sheet_name, display_label), ...] label 优先取标题日期,
    其次用 sheet 名. 兼容 cscan 调用方.
    """
    wb = load_workbook(str(xlsx_path), data_only=True)
    date_pattern = re.compile(r'^\d{1,2}\.\d{1,2}')
    results = []

    # 1. 日期格式 sheet
    date_sheets = [n for n in wb.sheetnames if date_pattern.match(n)]
    if date_sheets:
        for name in date_sheets:
            label = _extract_date_from_title(wb[name]) or name
            results.append((name, label))
        return results

    # 2. 兜底: 扫描所有 sheet (跳过 Sheet1)
    for name in wb.sheetnames:
        if name == "Sheet1":
            continue
        ws = wb[name]
        has_data = False
        for row in range(2, min(ws.max_row + 1, 100)):
            v = ws.cell(row, 1).value
            if v is not None and str(v).strip().isdigit():
                has_data = True
                break
        if not has_data:
            continue
        label = _extract_date_from_title(ws) or name
        results.append((name, label))
    return results


def find_kuanhouban_sheets(xlsx_path: str | Path) -> list[tuple[str, str]]:
    """找出宽厚板厂所有有数据的 sheet, 返回 [(sheet_name, date_label), ...].

    日期从标题 Row 1 提取 (如 "2026-4-21" → "4.21"), 取不到则用 sheet 名.
    跳过 Sheet1.
    """
    wb = load_workbook(str(xlsx_path), data_only=True)
    result = []
    for name in wb.sheetnames:
        if name == "Sheet1":
            continue
        ws = wb[name]
        # 检测是否有数据
        has_data = False
        for row in range(2, min(ws.max_row + 1, 100)):
            v = ws.cell(row, 1).value
            if v is not None and str(v).strip().isdigit():
                has_data = True
                break
        if not has_data:
            continue
        label = _extract_date_from_title(ws) or name
        result.append((name, label))
    return result


def extract_cscan_from_xlsx(
    xlsx_path: str | Path,
    images_dir: str | Path,
    sheet_name: str = "",
) -> dict[int, dict[str, str]]:
    """从 cscan xlsx 的指定 sheet 提取 F/G 列原图, 切出 8 个子图/行.

    sheet_name: 日期格式的 sheet 名 (如 "5.1"). 为空时自动选第一个日期 sheet.

    Returns:
        {row_idx: {"F_table": path, ...}}
    """
    xlsx_path = Path(xlsx_path)
    images_dir = Path(images_dir)
    images_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(str(xlsx_path), data_only=True)
    # 选 sheet 优先级: 传入名 > Sheet2 > 日期格式 > 最后一个
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "Sheet2" in wb.sheetnames:
        ws = wb["Sheet2"]
    else:
        sheets = find_cscan_sheets(xlsx_path)
        if sheets:
            ws = wb[sheets[0]]
        elif len(wb.sheetnames) > 1:
            ws = wb[wb.sheetnames[-1]]
        else:
            ws = wb[wb.sheetnames[0]]

    # 收集 F/G 列的所有图, 按 (row, col, sort) 排序
    targets = []
    for img_idx, img_obj in enumerate(ws._images):
        anchor = img_obj.anchor
        if not hasattr(anchor, "_from"):
            continue
        col_idx = anchor._from.col
        if col_idx not in (5, 6):     # F=5, G=6
            continue
        row_idx = anchor._from.row + 1   # 0-based → 1-based
        targets.append((row_idx, col_idx, img_idx, img_obj))

    targets.sort(key=lambda t: (t[0], t[1], t[2]))

    # 按行分组, 切图保存
    result: dict[int, dict[str, str]] = {}
    for row_idx, col_idx, img_idx, img_obj in targets:
        prefix = "F" if col_idx == 5 else "G"
        # 拿到原图字节, 写到临时文件 (view_splitter 需要文件路径)
        raw = img_obj._data()
        tmp_path = images_dir / f"_tmp_{prefix}_{row_idx:03d}_{img_idx}.png"
        tmp_path.write_bytes(raw)

        views = split_cscan_views(tmp_path, images_dir, prefix=prefix)

        # 保存分割结果
        row_dict = result.setdefault(row_idx, {})
        for en_name, src_path in views.items():
            if en_name == "warnings":
                continue
            src = Path(src_path)
            new_name = f"{prefix}_{en_name}-{row_idx:03d}{src.suffix}"
            new_path = images_dir / new_name
            if src != new_path:
                src.rename(new_path)
            row_dict[f"{prefix}_{en_name}"] = str(new_path)

        # 原图也保留 (用于板信息 OCR 等)
        full_name = f"{prefix}_full-{row_idx:03d}.png"
        full_path = images_dir / full_name
        tmp_path.rename(full_path)
        row_dict[f"{prefix}_full"] = str(full_path)

    return result


# ============================================================
# OCR: 13 列缺陷表格
# ============================================================
def ocr_defect_table(table_image_path: str | Path) -> list[dict[str, Any]]:
    """对切出的 13 列缺陷表格子图跑 pytesseract, 提取行数据.

    Returns:
        [
            {"序号": 18, "X起始": 868.0, "X终止": 971.0, ..., "幅值": 100.0},
            ...
        ]
    """
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        return []

    table_image_path = Path(table_image_path)
    if not table_image_path.exists():
        return []

    img = Image.open(table_image_path)
    w, h = img.size
    # 放大至 ~3000px 宽, 确保 13 列每列 > 200px
    scale = max(3, int(3000 / max(w, 1)))
    img = img.resize((w * scale, h * scale), Image.LANCZOS)

    # 彩色表格 → 灰度 → 自适应二值化 (白底黑字)
    try:
        import cv2, numpy as np
        arr = np.array(img)
        if arr.ndim == 3:
            gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
        else:
            gray = np.array(img)
        # Otsu 自适应阈值: 字暗背景亮 → THRESH_BINARY 保持白底黑字
        _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        from PIL import Image as _PILImage
        img = _PILImage.fromarray(bw, mode='L')
    except Exception:
        img = img.convert('L')

    try:
        data = pytesseract.image_to_data(
            img,
            lang="chi_sim+eng",
            output_type=pytesseract.Output.DICT,
            config="--psm 6",
        )
    except Exception as e:
        return [{"error": f"OCR 失败: {e}"}]

    # 收集所有 numeric token (去掉汉字标签)
    items = []
    for i in range(len(data["text"])):
        text = str(data["text"][i]).strip()
        if not text:
            continue
        try:
            conf = float(data["conf"][i])
        except (ValueError, TypeError):
            conf = -1
        if conf < 15:  # 降低: 红底单元格 conf 可能偏低
            continue
        t = text.replace(",", "").replace("，", "")
        try:
            float(t)
            items.append({
                "text": text,
                "value": float(t) if "." in t else int(t),
                "left": data["left"][i],
                "top": data["top"][i],
            })
        except ValueError:
            continue

    if not items:
        from .llm_ocr import _image_has_content
        if _image_has_content(table_image_path):
            return [{"warning": "OCR 识别失败，图片有内容但未读出"}]
        return []

    # 算法: 从 items 中找出每一条数据行.
    # 关键观察: 表头行在最上面 (Y 起始值 = 2733.3 之类), 实际数据行在表头下面.
    # 表里 "Y 起始/终止/中点" 列的值集中在头几列有大量相同值 (2733.3 等), 用于识别表头.
    # 用更稳的策略: 按 top 聚类, 然后每行按 left 排序, 找最长的那个数组 (≥13 个数字).
    items.sort(key=lambda x: (x["top"], x["left"]))

    # 按 top 间隔聚类 (60 px 容差, 表格行高约 60 px)
    clusters = []
    for it in items:
        if clusters and abs(clusters[-1][0]["top"] - it["top"]) < 60:
            clusters[-1].append(it)
        else:
            clusters.append([it])

    # 找每行 cluster 中前 13 个数值
    results = []
    for cluster in clusters:
        # 按 left 排序 (行内的列顺序)
        cluster.sort(key=lambda x: x["left"])
        # 去重 (OCR 偶尔把同一数字拆成两个 token, left < 5 px 视为重复)
        deduped = []
        for it in cluster:
            if deduped and abs(it["left"] - deduped[-1]["left"]) < 5:
                continue
            deduped.append(it)
        vals = [it["value"] for it in deduped]
        if len(vals) < 10:  # 放宽: 部分行 OCR 丢 1-3 个 token 仍保留
            continue
        nums = vals[:13]
        row = {"序号": int(nums[0])}
        for col_name, val in zip(DEFECT_TABLE_COLS[1:], nums[1:13]):
            row[col_name] = val
        results.append(row)

    return results


# ============================================================
# OCR: 板信息 panel
# ============================================================
BOARD_LABELS = {
    "板号": "plate_no",
    "探伤代号": "test_code",
    "钢种": "grade",
    "生产日期": "prod_date",
    "检测日期": "test_date",
    "标准号": "standard",
    "厚度": "thickness",
    "长度": "length",
    "宽度": "width",
}


def ocr_board_info(board_image_path: str | Path) -> dict[str, Any]:
    """识别板信息 panel. 裁剪原图右下角面板 (x>68%%, y>68%%), 放大 4x + 二值化.

    板信息 panel 有白底网格线, 含: 板号/探伤代号/钢种/生产日期/检测日期/标准号/厚度/长度/宽度
    Tesseract 精度有限, 数值可能有 OCR 误差; 宁可返回近似值也不要 None.
    """
    try:
        import pytesseract
        import cv2
        import re as _re
    except ImportError:
        return {}

    board_image_path = Path(board_image_path)
    if not board_image_path.exists():
        return {}

    img_cv = cv2.imread(str(board_image_path))
    if img_cv is None:
        return {}
    h, w = img_cv.shape[:2]

    # 右下角 ~30% 区域 (板信息面板)
    x0, y0 = int(w * 0.65), int(h * 0.65)
    crop = img_cv[y0:, x0:]
    # 放大 4x + 二值化反相 (白底黑字更适合 Tesseract)
    big = cv2.resize(crop, None, fx=4, fy=4, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(big, cv2.COLOR_BGR2GRAY)
    _, bw = cv2.threshold(gray, 140, 255, cv2.THRESH_BINARY)
    # 白底黑字 → 直接用 image_to_string (不用 image_to_data, 数据少时更稳)
    text = pytesseract.image_to_string(bw, lang="chi_sim+eng", config="--psm 6")

    result: dict[str, Any] = {}
    LABEL_PATTERNS = [
        (r"板\s*号?\s*[::]?\s*(\d{10,16})", "plate_no"),
        (r"探伤\s*代?\s*码?\s*[::]?\s*(\S+)", "test_code"),
        (r"钢\s*种\s*[::]?\s*(\S+)", "grade"),
        (r"生产\s*日?\s*期?\s*[::]?\s*(\d{8})", "prod_date"),
        (r"检测\s*日?\s*期?\s*[::]?\s*(\S+)", "test_date"),
        (r"标\s*准\s*号?\s*[::]?\s*(\S+)", "standard"),
        (r"厚\s*度?\s*[\[\(]?m*m*[\]\)]?\s*[::]?\s*(\d+\.?\d*)", "thickness"),
        (r"长\s*度?\s*[\[\(]?m*m*[\]\)]?\s*[::]?\s*(\d+\.?\d*)", "length"),
        (r"宽\s*度?\s*[\[\(]?m*m*[\]\)]?\s*[::]?\s*(\d+\.?\d*)", "width"),
    ]
    for pat, key in LABEL_PATTERNS:
        m = _re.search(pat, text)
        if m:
            val = m.group(1)
            try:
                if "." in val: val = float(val)
                else: val = int(val)
            except ValueError:
                pass
            result[key] = val
    return result

