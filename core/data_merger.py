"""数据整合模块。

将所有处理结果合并成统一的缺陷记录数据结构，并保存为 JSON。

数据流:
    1. importer 读取 xls/xlsx 中的文字字段 -> records[]
    2. image_extractor 提取图片 -> images[]
    3. view_splitter 切分图片为三视图 -> views[]
    4. ocr_extractor 识别图片中的参数 -> ocr_results[]
    5. data_merger 将以上所有信息按 row_index 关联 -> defect_records.json
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any


class DataMerger:
    """数据整合器。

    将 importer / image_extractor / view_splitter / ocr_extractor 的输出
    按 row_index 关联为最终的缺陷记录。
    """

    def __init__(self, output_dir: str | Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def merge(
        self,
        records: list[dict],
        images: list[dict],
        views: list[dict] | None = None,
        ocr_results: list[dict] | None = None,
    ) -> list[dict]:
        """合并所有数据。

        Parameters
        ----------
        records : list[dict]
            importer 返回的缺陷记录 (含 row_index, 钢种, 类别 等)
        images : list[dict]
            image_extractor 返回的图片信息 (含 row_index, file_path, format)
        views : list[dict] | None
            view_splitter 返回的视图信息 (含 source, views 字典)
        ocr_results : list[dict] | None
            ocr_extractor 返回的识别结果 (含 source, params)

        Returns
        -------
        list[dict]
            合并后的缺陷记录, 每条记录的字段:
            {
                # 基础字段 (来自 importer)
                "row_index": int,
                "序号": str,
                "生产厂": str,
                "钢板号": str,
                "钢种": str,
                "类别": str,
                "缺陷分析": str,

                # 图片 (来自 image_extractor)
                "图-1": str | None,           # 原图1 路径
                "图-2": str | None,           # 原图2 路径
                "图-1_format": str,
                "图-2_format": str,

                # 视图 (来自 view_splitter)
                "俯视图-1": str | None,
                "长边方向侧视图-1": str | None,
                "短边方向侧视图-1": str | None,
                "俯视图-2": str | None,
                "长边方向侧视图-2": str | None,
                "短边方向侧视图-2": str | None,
                "视图标注预览-1": str | None,
                "视图标注预览-2": str | None,

                # 缺陷参数 (来自 ocr_extractor)
                "缺陷数据": {
                    "材料尺寸": "...",
                    "缺陷中心X": "...",
                    "缺陷中心Y": "...",
                    "缺陷长度": "...",
                    "缺陷宽度": "...",
                    "缺陷深度": "...",
                },
                "ocr_raw_text": str,

                # 警告
                "warnings": list[str],
            }
        """
        views = views or []
        ocr_results = ocr_results or []

        # 按 row_index 索引图片
        images_by_row: dict[int, list[dict]] = {}
        for img in images:
            row = img.get("row_index")
            if row is None:
                continue
            images_by_row.setdefault(row, []).append(img)

        # 按 source (图片路径) 索引视图和 OCR 结果
        views_by_source: dict[str, dict] = {v.get("source", ""): v for v in views}
        ocr_by_source: dict[str, dict] = {
            o.get("source", ""): o for o in ocr_results
        }

        # 合并
        merged_records = []
        for rec in records:
            row_idx = rec.get("row_index")
            row_images = images_by_row.get(row_idx, [])

            # 初始化合并字段
            merged = dict(rec)
            merged.update(
                {
                    "图-1": None,
                    "图-2": None,
                    "图-1_format": None,
                    "图-2_format": None,
                    "俯视图-1": None,
                    "长边方向侧视图-1": None,
                    "短边方向侧视图-1": None,
                    "俯视图-2": None,
                    "长边方向侧视图-2": None,
                    "短边方向侧视图-2": None,
                    "视图标注预览-1": None,
                    "视图标注预览-2": None,
                    "缺陷数据": {},
                    "ocr_raw_text": "",
                    "warnings": [],
                }
            )

            # 关联图片
            warnings = []
            for idx, img in enumerate(row_images[:2]):
                img_num = idx + 1
                merged[f"图-{img_num}"] = img.get("file_path")
                merged[f"图-{img_num}_format"] = img.get("format")

                # 关联视图
                img_path = img.get("file_path", "")
                if img_path in views_by_source:
                    view_result = views_by_source[img_path]
                    view_dict = view_result.get("views", {})
                    merged[f"俯视图-{img_num}"] = view_dict.get("1-俯视图")
                    merged[f"长边方向侧视图-{img_num}"] = view_dict.get("2-长边侧视图")
                    merged[f"短边方向侧视图-{img_num}"] = view_dict.get("3-短边侧视图")
                    merged[f"视图标注预览-{img_num}"] = view_result.get("annotated")

                # 关联 OCR (同一行只 OCR 第一张图, 因为两张图信息类似)
                # OCR 跑的是切图 + 原图, 通过原图路径或 stem 匹配
                if idx == 0:
                    ocr = None
                    if img_path in ocr_by_source:
                        ocr = ocr_by_source[img_path]
                    else:
                        img_stem = Path(img_path).stem
                        for src, o in ocr_by_source.items():
                            if img_stem in src:
                                ocr = o
                                break

                    if ocr is not None:
                        merged["缺陷数据"] = ocr.get("params", {})
                        merged["ocr_raw_text"] = ocr.get("full_text", "")
                        if ocr.get("warnings"):
                            warnings.extend(ocr["warnings"])

            # 视图警告
            if row_images and not any(
                merged.get(f"俯视图-{i}") for i in (1, 2) if i <= len(row_images)
            ):
                warnings.append("未能切分出三视图")

            if not row_images:
                warnings.append("该行没有图片")

            merged["warnings"] = warnings
            merged_records.append(merged)

        return merged_records

    def save_json(
        self,
        records: list[dict],
        file_name: str = "defect_records.json",
    ) -> Path:
        """保存为 JSON。"""
        json_path = self.output_dir / file_name
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        return json_path

    # Excel 中图片列名 -> 记录字段名的映射
    _IMAGE_COLUMNS = {
        "图-1",
        "图-2",
        "俯视图-1",
        "长边方向侧视图-1",
        "短边方向侧视图-1",
        "俯视图-2",
        "长边方向侧视图-2",
        "短边方向侧视图-2",
    }

    # 空图片列的默认行高 (磅), 用于无图的行让版式不至于太扁
    _EMPTY_IMAGE_ROW_HEIGHT = 15

    # 图片在 xlsx 中显示的缩放比例 (相对于原始像素).
    # 原图字节不变, 只是 openpyxl 写入的显示尺寸缩小; 实际 PNG 仍是 1920×1080 等.
    # ¼ → 1920×1080 显示为 480×270, 单元格不至于撑爆.
    _DISPLAY_SCALE = 0.25

    def _embed_image(self, ws, cell, img_path: str):
        """将图片嵌入到指定单元格。

        原图像素不变, 仅按 _DISPLAY_SCALE 缩小在 xlsx 中显示的尺寸 (列宽 / 行高同步).
        失败时 (路径不存在 / 不支持的格式) 在单元格内写回原路径字符串作为回退。
        """
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.units import pixels_to_EMU

        if not img_path:
            return

        path = Path(img_path)
        if not path.is_file():
            # 图片缺失, 退化为文本提示
            ws.cell(row=cell.row, column=cell.column, value=f"[图片缺失] {img_path}")
            return

        try:
            img = XLImage(str(path))
        except Exception as e:
            ws.cell(row=cell.row, column=cell.column, value=f"[图片加载失败] {img_path} ({e})")
            return

        # 原图原始像素
        raw_w = img.width or 0
        raw_h = img.height or 0
        # xlsx 中显示尺寸 (按 _DISPLAY_SCALE 缩小)
        disp_w = int(raw_w * self._DISPLAY_SCALE) if raw_w else 0
        disp_h = int(raw_h * self._DISPLAY_SCALE) if raw_h else 0
        img.width = disp_w
        img.height = disp_h

        # 锚定到单元格
        img.anchor = cell.coordinate
        ws.add_image(img)

        # openpyxl 默认用 OneCellAnchor; 其 ext 决定显示尺寸 (EMU 单位).
        # 必须同步设 ext, 否则 Excel 仍按原图 1920×1080 显示.
        if disp_w and disp_h and hasattr(img.anchor, 'ext') and img.anchor.ext is not None:
            img.anchor.ext.cx = pixels_to_EMU(disp_w)
            img.anchor.ext.cy = pixels_to_EMU(disp_h)

        # 列宽: openpyxl 列宽单位是字符宽度, 1 px ≈ 0.14 个字符
        col_letter = get_column_letter(cell.column)
        if disp_w:
            target_col_width = int(disp_w * 0.14) + 2
        else:
            target_col_width = 15
        current_col_width = ws.column_dimensions[col_letter].width or 0
        ws.column_dimensions[col_letter].width = max(current_col_width, target_col_width)

        # 行高: openpyxl 行高单位是磅, 1 px ≈ 0.75 pt
        if disp_h:
            target_row_height = int(disp_h * 0.75) + 4
        else:
            target_row_height = self._EMPTY_IMAGE_ROW_HEIGHT
        current_row_height = ws.row_dimensions[cell.row].height or 0
        ws.row_dimensions[cell.row].height = max(current_row_height, target_row_height)

    def save_excel(self, records: list[dict]) -> Path:
        """保存为 Excel 文件, 方便用户用 Excel 查看。

        图片列(图-1/图-2/俯视图/侧视图等)直接嵌入实际图片,
        而非保存路径字符串, 方便用户在 Excel 中直接预览。
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("需要 openpyxl: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "缺陷记录"

        # 表头
        headers = [
            "序号",
            "生产厂",
            "钢板号",
            "钢种",
            "类别",
            "缺陷分析",
            "图-1",
            "图-2",
            "俯视图-1",
            "长边方向侧视图-1",
            "短边方向侧视图-1",
            "俯视图-2",
            "长边方向侧视图-2",
            "短边方向侧视图-2",
            "材料尺寸",
            "缺陷中心X",
            "缺陷中心Y",
            "缺陷长度",
            "缺陷宽度",
            "缺陷深度",
            "缺陷面积",
            "C扫描值",
            "警告",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 数据行
        for row_idx, rec in enumerate(records, 2):
            params = rec.get("缺陷数据", {})
            row_data = [
                rec.get("序号", ""),
                rec.get("生产厂", ""),
                rec.get("钢板号", ""),
                rec.get("钢种", ""),
                rec.get("类别", ""),
                rec.get("缺陷分析", ""),
                rec.get("图-1", ""),
                rec.get("图-2", ""),
                rec.get("俯视图-1", ""),
                rec.get("长边方向侧视图-1", ""),
                rec.get("短边方向侧视图-1", ""),
                rec.get("俯视图-2", ""),
                rec.get("长边方向侧视图-2", ""),
                rec.get("短边方向侧视图-2", ""),
                params.get("材料尺寸", ""),
                params.get("缺陷中心X", ""),
                params.get("缺陷中心Y", ""),
                params.get("缺陷长度", ""),
                params.get("缺陷宽度", ""),
                params.get("缺陷深度", ""),
                params.get("缺陷面积", ""),
                params.get("C扫描值", ""),
                "; ".join(rec.get("warnings", [])),
            ]
            for col_idx, val in enumerate(row_data, 1):
                header = headers[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                # 图片列: 先占位空单元格, 写入文本会导致图片锚定失效,
                # 因此统一由 _embed_image 处理 (含空路径场景)
                if header in self._IMAGE_COLUMNS:
                    if val:
                        self._embed_image(ws, cell, val)
                    else:
                        # 保留空单元格, 但设置行高让无图行不显瘦
                        ws.row_dimensions[row_idx].height = max(
                            ws.row_dimensions[row_idx].height or 0,
                            self._EMPTY_IMAGE_ROW_HEIGHT,
                        )
                else:
                    cell.value = val

        # 列宽自适应 (只针对非图片列, 图片列已在 _embed_image 中按图片尺寸调整)
        from openpyxl.utils import get_column_letter

        for col_idx, h in enumerate(headers, 1):
            if h in self._IMAGE_COLUMNS:
                continue
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(
                15, min(40, len(h) * 2 + 5)
            )

        excel_path = self.output_dir / "defect_records.xlsx"
        wb.save(str(excel_path))
        return excel_path


def merge_and_save(
    records: list[dict],
    images: list[dict],
    output_dir: str | Path,
    views: list[dict] | None = None,
    ocr_results: list[dict] | None = None,
) -> tuple[list[dict], Path, Path]:
    """便捷函数：合并并保存为 JSON 和 Excel。"""
    merger = DataMerger(output_dir)
    merged = merger.merge(records, images, views, ocr_results)
    json_path = merger.save_json(merged)
    excel_path = merger.save_excel(merged)
    return merged, json_path, excel_path


if __name__ == "__main__":
    # 测试合并
    sample_records = [
        {
            "row_index": 7,
            "序号": "1.0",
            "钢板号": "26102421340201",
            "钢种": "JX22145-2025 X60MS",
            "类别": "管线",
            "缺陷分析": "显孔",
        }
    ]
    sample_images = [
        {
            "row_index": 7,
            "column": "缺陷图谱",
            "image_index": 1,
            "file_path": "/tmp/img1.png",
            "format": "png",
        },
        {
            "row_index": 7,
            "column": "缺陷图谱",
            "image_index": 2,
            "file_path": "/tmp/img2.png",
            "format": "png",
        },
    ]
    merged, json_path, excel_path = merge_and_save(
        sample_records, sample_images, "./test_output"
    )
    print(f"JSON: {json_path}")
    print(f"Excel: {excel_path}")
    print(f"\n合并结果示例:")
    print(json.dumps(merged[0], ensure_ascii=False, indent=2))