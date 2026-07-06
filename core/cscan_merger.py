"""cscan 文档的数据合并 (中厚板卷厂).

- merge_cscan_records: 从 xlsx 表格行 + 子图路径 + (后续 OCR 结果) 打包成 cscan_records 列表
- save_json: 写到 output/<task_id>/cscan_records.json
- save_to_db: 写到 SQLite cscan_records 表
"""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# cscan_records.json 每条的 schema.
# OCR 字段 (缺陷表格/板信息) 暂时留 None 或空, 后续 Task 1.x 补 OCR 时填.
CSCAN_RECORD_FIELDS = (
    "序号", "生产厂", "钢板号", "钢种", "类别", "缺陷分析",
    # 子图路径 (8 张)
    "F_table", "F_ascan", "F_cscan",
    "G_table", "G_ascan", "G_cscan",
    # OCR 字段
    "缺陷表格_F", "缺陷表格_G",   # list[dict], 13 列每行
    "板号", "探伤代号", "钢种OCR", "生产日期", "检测日期",
    "标准号", "厚度", "长度", "宽度", "warnings",
)


def merge_cscan_records(
    xlsx_path: str | Path,
    image_map: dict[int, dict[str, str]],
    ocr_table_map: dict[int, dict[str, list[dict[str, Any]]]] | None = None,
    ocr_board_map: dict[int, dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """从 xlsx 的 5.1 sheet 提取基础字段, 与 image_map/OCR 结果合并.

    Args:
        xlsx_path: cscan xlsx 路径 (用 5.1 sheet, 索引 1)
        image_map: {row_idx: {"F_table": path, ...}} — cscan_ocr.extract_cscan_from_xlsx 的输出
        ocr_table_map: {row_idx: list[dict]} — 缺陷表格 OCR 结果 (可选)
        image_map: {row_idx: {F_table: path, ...}}

    Returns:
        list of cscan record dicts
    """
    xlsx_path = Path(xlsx_path)
    ocr_table_map = ocr_table_map or {}
    ocr_board_map = ocr_board_map or {}

    wb = load_workbook(str(xlsx_path), data_only=True)
    if len(wb.sheetnames) < 2:
        raise ValueError(
            f"cscan 文件至少需要 2 个 sheet, 当前 {len(wb.sheetnames)} 个: {wb.sheetnames}"
        )
    ws = wb[wb.sheetnames[1]]   # 5.1

    # 5.1 行 7+ 是真实数据 (行 1=标题, 2=表头, 3-4=要求, 5=示意, 6=空)
    DATA_START_ROW = 7

    records = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        # 从 xlsx 取基础 8 列
        序号 = ws.cell(row_idx, 1).value
        if 序号 is None or 序号 == "":
            continue  # 跳过空行
        # column 2-8: 生产厂/钢板号/钢种/类别/缺陷图谱/缺陷照片/缺陷分析
        钢板号 = str(ws.cell(row_idx, 3).value or "").strip()
        钢种 = str(ws.cell(row_idx, 4).value or "").strip()
        if not 钢板号 and not 钢种:
            continue  # 钢板号和钢种都空 → 跳过
        record: dict[str, Any] = {
            "row_index": row_idx + 1,
            "序号": str(序号),
            "生产厂": str(ws.cell(row_idx, 2).value or "").strip(),
            "钢板号": 钢板号,
            "钢种": 钢种,
            "类别": str(ws.cell(row_idx, 5).value or "").strip(),
            "缺陷分析": str(ws.cell(row_idx, 8).value or "").strip(),
        }
        # 子图路径
        record.update(image_map.get(row_idx, {}))
        # 缺的子图设为 None (跨 F/G 列缺失的情况, 比如行只有 F 列图)
        for k in CSCAN_RECORD_FIELDS:
            if k.startswith(("F_", "G_")) and k not in record:
                record[k] = None
        # OCR 字段 — F/G 表格分别存
        tables = ocr_table_map.get(row_idx, {})
        record["缺陷表格_F"] = tables.get("F", [])
        record["缺陷表格_G"] = tables.get("G", [])
        # 板信息字段: 从 OCR 结果获取
        board = ocr_board_map.get(row_idx, {})
        for en, cn in [
            ("plate_no", "板号"), ("test_code", "探伤代号"),
            ("grade", "钢种OCR"), ("prod_date", "生产日期"),
            ("test_date", "检测日期"), ("standard", "标准号"),
            ("thickness", "厚度"), ("length", "长度"), ("width", "宽度"),
        ]:
            record[cn] = board.get(en)
        record["warnings"] = []
        records.append(record)

    return records


# ============================================================
# xlsx 导出 (openpyxl, 参考 data_merger.save_excel)
# ============================================================
_CSCAN_IMAGE_COLUMNS = {
    "F_table", "F_ascan", "F_cscan",
    "G_table", "G_ascan", "G_cscan",
}
_DISPLAY_SCALE = 0.25   # 图片在 xlsx 中显示缩放 (与 data_merger 一致)


def save_excel(records: list[dict[str, Any]], output_dir: str | Path) -> Path:
    """生成 cscan_records.xlsx (两个 sheet).

    Sheet 1 "缺陷记录": 基础字段 + 板信息 + 6 张子图
    Sheet 2 "缺陷表格": 每条记录的 13 列缺陷明细
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.units import pixels_to_EMU
    except ImportError:
        raise ImportError("需要 openpyxl")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ==================== Sheet 1: 缺陷记录 ====================
    ws1 = wb.active
    ws1.title = "缺陷记录"
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_align = Alignment(horizontal="center")

    # 表头: 文本列 + 图片列
    text_headers = [
        "序号", "生产厂", "钢板号", "钢种", "类别", "缺陷分析",
        "厚度[mm]", "长度[mm]", "宽度[mm]",
        "板号", "探伤代号", "钢种OCR", "生产日期", "检测日期", "标准号",
        "warnings",
    ]
    img_headers = ["F_table", "F_ascan", "F_cscan", "G_table", "G_ascan", "G_cscan"]
    all_headers = text_headers + img_headers

    for col_idx, h in enumerate(all_headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, rec in enumerate(records, 2):
        # 文本列
        text_vals = [
            rec.get("序号", ""), rec.get("生产厂", ""), rec.get("钢板号", ""),
            rec.get("钢种", ""), rec.get("类别", ""), rec.get("缺陷分析", ""),
            rec.get("厚度", ""), rec.get("长度", ""), rec.get("宽度", ""),
            rec.get("板号", ""), rec.get("探伤代号", ""),
            rec.get("钢种OCR", ""), rec.get("生产日期", ""),
            rec.get("检测日期", ""), rec.get("标准号", ""),
            "; ".join(rec.get("warnings", [])) if isinstance(rec.get("warnings"), list) else str(rec.get("warnings", "") or ""),
        ]
        for ci, val in enumerate(text_vals, 1):
            cell = ws1.cell(row=row_idx, column=ci, value=val if val is not None else "")

        # 图片列
        for ci_offset, h in enumerate(img_headers):
            col_idx = len(text_headers) + ci_offset + 1
            val = rec.get(h, "")
            cell = ws1.cell(row=row_idx, column=col_idx, value="")
            if val and isinstance(val, str):
                path = Path(val)
                if path.is_file():
                    try:
                        img = XLImage(str(path))
                        raw_w, raw_h = img.width or 0, img.height or 0
                        disp_w, disp_h = int(raw_w * _DISPLAY_SCALE), int(raw_h * _DISPLAY_SCALE)
                        img.width, img.height = disp_w, disp_h
                        img.anchor = cell.coordinate
                        ws1.add_image(img)
                        if hasattr(img.anchor, 'ext') and img.anchor.ext is not None and disp_w and disp_h:
                            img.anchor.ext.cx = pixels_to_EMU(disp_w)
                            img.anchor.ext.cy = pixels_to_EMU(disp_h)
                        cl = get_column_letter(cell.column)
                        ws1.column_dimensions[cl].width = max(
                            ws1.column_dimensions[cl].width or 10, int(disp_w * 0.14) + 2
                        )
                        ws1.row_dimensions[cell.row].height = max(
                            ws1.row_dimensions[cell.row].height or 10, int(disp_h * 0.75) + 4
                        )
                        continue
                    except Exception:
                        pass
                cell.value = val if val else ""

    # 文本列宽自适应
    for ci in range(1, len(text_headers) + 1):
        cl = get_column_letter(ci)
        ws1.column_dimensions[cl].width = max(12, min(40, len(all_headers[ci - 1]) * 2 + 3))

    # ==================== Sheet 2+3: F/G 缺陷表格 ====================
    defect_cols = ["钢板号", "序号", "X起始", "X终止", "X中点", "X长度",
                   "Y起始", "Y终止", "Y中点", "Y长度", "面积", "类型", "深度", "幅值"]
    for suffix in ("F", "G"):
        ws2 = wb.create_sheet(f"缺陷表格_{suffix}")
        for col_idx, h in enumerate(defect_cols, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        sheet_row = 2
        for rec in records:
            plate = rec.get("钢板号", "")
            dt = rec.get(f"缺陷表格_{suffix}") or []
            for drow in dt:
                if isinstance(drow, list):
                    vals = [plate] + drow
                else:
                    vals = [plate] + [drow.get(c, "") for c in defect_cols[1:]]
                for ci, val in enumerate(vals, 1):
                    ws2.cell(row=sheet_row, column=ci, value=val if val is not None else "")
                sheet_row += 1
        for ci in range(1, len(defect_cols) + 1):
            cl = get_column_letter(ci)
            ws2.column_dimensions[cl].width = max(10, min(20, len(defect_cols[ci - 1]) * 2 + 5))

    excel_path = output_dir / "cscan_records.xlsx"
    wb.save(str(excel_path))
    return excel_path


def save_json(records: list[dict[str, Any]], output_dir: str | Path) -> Path:
    """写到 output_dir/cscan_records.json."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "cscan_records.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({"count": len(records), "records": records}, f, ensure_ascii=False, indent=2)
    return json_path


# ============================================================
# 数据库 (SQLite)
# ============================================================
def ensure_cscan_table(conn: sqlite3.Connection) -> None:
    """确保 cscan_records 表存在."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cscan_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            "序号" TEXT, "生产厂" TEXT, "钢板号" TEXT, "钢种" TEXT,
            "类别" TEXT, "缺陷分析" TEXT,
            F_table TEXT, F_ascan TEXT, F_cscan TEXT,
            G_table TEXT, G_ascan TEXT, G_cscan TEXT,
            缺陷表格_F TEXT, 缺陷表格_G TEXT,
            "板号" TEXT, "探伤代号" TEXT, "钢种OCR" TEXT, "生产日期" TEXT,
            "检测日期" TEXT, "标准号" TEXT, "厚度" REAL, "长度" REAL, "宽度" REAL,
            warnings TEXT,
            created_at REAL,
            UNIQUE(task_id, row_index)
        )
    """)
    conn.commit()


def save_to_db(conn: sqlite3.Connection, task_id: str, records: list[dict[str, Any]]) -> int:
    """写入 SQLite, 返回插入/更新的行数."""
    ensure_cscan_table(conn)
    import time
    now = time.time()
    count = 0
    for r in records:
        # 缺陷表格 F/G 分别序列化为 JSON
        defect_table_f_json = json.dumps(r.get("缺陷表格_F") or [], ensure_ascii=False)
        defect_table_g_json = json.dumps(r.get("缺陷表格_G") or [], ensure_ascii=False)
        warnings_json = json.dumps(r.get("warnings") or [], ensure_ascii=False)
        row_1based = r.get("row_index", 0)
        conn.execute("""
            INSERT OR REPLACE INTO cscan_records
            (task_id, row_index, "序号", "生产厂", "钢板号", "钢种", "类别", "缺陷分析",
             F_table, F_ascan, F_cscan,
             G_table, G_ascan, G_cscan,
             缺陷表格_F, 缺陷表格_G, "板号", "探伤代号", "钢种OCR", "生产日期", "检测日期",
             "标准号", "厚度", "长度", "宽度", warnings, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            task_id, row_1based,
            r.get("序号"), r.get("生产厂"), r.get("钢板号"), r.get("钢种"),
            r.get("类别"), r.get("缺陷分析"),
            r.get("F_table"), r.get("F_ascan"), r.get("F_cscan"),
            r.get("G_table"), r.get("G_ascan"), r.get("G_cscan"),
            defect_table_f_json, defect_table_g_json,
            r.get("板号"), r.get("探伤代号"), r.get("钢种OCR"), r.get("生产日期"),
            r.get("检测日期"), r.get("标准号"),
            r.get("厚度"), r.get("长度"), r.get("宽度"),
            warnings_json, now,
        ))
        count += 1
    conn.commit()
    return count


def load_from_db(conn: sqlite3.Connection, task_id: str) -> list[dict[str, Any]]:
    """从 SQLite 读 cscan_records."""
    ensure_cscan_table(conn)
    rows = conn.execute(
        "SELECT * FROM cscan_records WHERE task_id = ? ORDER BY row_index",
        (task_id,),
    ).fetchall()
    cols = [d[0] for d in conn.execute("PRAGMA table_info(cscan_records)").fetchall()]
    results = []
    for row in rows:
        d = dict(zip(cols, row))
        # 反序列化 JSON
        for fld in ("缺陷表格_F", "缺陷表格_G"):
            try:
                d[fld] = json.loads(d.get(fld) or "[]")
            except Exception:
                d[fld] = []
        try:
            d["warnings"] = json.loads(d.get("warnings") or "[]")
        except Exception:
            d["warnings"] = []
        results.append(d)
    return results
